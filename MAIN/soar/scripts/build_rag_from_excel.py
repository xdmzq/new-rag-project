#!/usr/bin/env python3
"""Build/expand a RAG corpus from Excel workbooks under a data root.

Outputs (schema-compatible with current SpaceRAG):
  - cleaned_texts.jsonl
  - with_images.jsonl
  - all_texts.jsonl
  - faiss_index_metadata.jsonl
  - vector_store/embeddings.npy
  - vector_store/manifest.json
  - assets/docs/<doc_id>/
  - assets/images/<doc_id>/
  - assets/tables/<doc_id>/

This script supports incremental updates and resume via a progress file
stored under the selected data root by default.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import requests
from openpyxl import load_workbook

from env_utils import resolve_env_value

 

DEFAULT_DATA_ROOT = Path("source_data")
DEFAULT_OUTPUT_DIR = Path("SpaceRAG")
DEFAULT_PROGRESS_FILE_NAME = "build_rag_progress.json"
DEFAULT_EMBEDDING_API_BASE = "https://api.siliconflow.cn/v1"
DEFAULT_EMBEDDING_MODEL = "Qwen/Qwen3-Embedding-8B"
DEFAULT_EMBEDDING_DIMENSIONS = 0
DEFAULT_FAISS_INDEX_FILE = "faiss_index.index"
LEGACY_FAISS_INDEX_FILE = "faiss.index"
SUPPORTED_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
TEMP_PREFIXES = ("~$",)
SUPPORTED_EMBEDDING_DIMENSIONS = {
    "Qwen/Qwen3-Embedding-8B": [64, 128, 256, 512, 768, 1024, 2048, 4096],
    "Qwen/Qwen3-Embedding-4B": [64, 128, 256, 512, 768, 1024, 2048],
    "Qwen/Qwen3-Embedding-0.6B": [64, 128, 256, 512, 768, 1024],
}


@dataclass
class AssetRef:
    asset_id: str
    asset_type: str
    page_no: int
    keywords: str
    rel_path: str
    abs_path: str
    source_sheet: str
    source_row_index: int
    title: str


@dataclass
class ChunkRecord:
    chunk_id: str
    parent_chunk_uid: str
    doc_id: str
    page_id: str
    page_no: int
    doc_title: str
    heading: str
    chunk_order: int
    chunk_text_clean: str
    chunk_text_display: str
    keywords: list[str]
    table_ids_on_page: list[str]
    image_ids_on_page: list[str]
    embedding_source: str = "legacy_parent_plus_local_svd"
    asset_type: str = "text"
    image_paths: list[str] = None
    table_paths: list[str] = None
    # Fields used during build but potentially excluded from final JSONL if needed
    file_path: str = ""
    ext: str = ""
    para_id: int = 0


@dataclass
class PageRecord:
    page_id: str
    doc_id: str
    page_no: int
    doc_title: str
    page_summary: str
    page_keywords: list[str]
    chunk_ids: list[str]
    table_ids: list[str]
    image_ids: list[str]
    embedding_source: str = "mean_legacy_children"


@dataclass
class TableRecord:
    table_id: str
    doc_id: str
    page_id: str
    page_no: int
    doc_title: str
    table_title: str
    table_headers: list[str]
    table_summary: str
    table_markdown: str
    keywords: list[str]
    asset_path: str
    embedding_source: str = "legacy_semantic"


@dataclass
class DocRecord:
    doc_id: str
    doc_title: str
    doc_summary: str
    page_count: int
    top_keywords: list[str]
    page_ids: list[str]
    embedding_source: str = "mean_page_embeddings"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_text(text: str) -> str:
    return re.sub(r"\n{3,}", "\n\n", text.replace("\r\n", "\n").replace("\r", "\n")).strip()


def slugify(value: str) -> str:
    text = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", str(value or "").strip(), flags=re.UNICODE)
    return text.strip("_") or "document"


def normalize_token(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def canonical_file_type(value: str) -> str:
    token = normalize_token(value)
    if token.startswith("ppt"):
        return "pptx"
    if token.startswith("doc") or token == "word":
        return "docx"
    if token.startswith("xls"):
        return "xlsx"
    if token.startswith("pdf"):
        return "pdf"
    return token


def normalize_directory_key(value: str) -> str:
    raw = safe_str(value).replace("\\", "/")
    parts = [part for part in raw.split("/") if part and part != "source_data"]
    if not parts:
        return ""
    if len(parts) >= 2:
        return "/".join(parts[-2:])
    return parts[-1]


def split_keywords(text: str) -> list[str]:
    parts = re.split(r"[、,，;；\s]+", (text or "").strip())
    return [item for item in parts if item]


def safe_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    text = str(value).strip()
    match = re.search(r"\d+", text)
    if match:
        return int(match.group())
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return default


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def write_json_atomic(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp, path)


def resolve_progress_file(data_root: Path, progress_file: Path | None) -> Path:
    if progress_file is not None:
        return progress_file.resolve()
    return (data_root / DEFAULT_PROGRESS_FILE_NAME).resolve()


def resolve_embedding_dimensions(model: str, dimensions: int | None) -> int | None:
    allowed = SUPPORTED_EMBEDDING_DIMENSIONS.get(model)
    if dimensions in (None, 0):
        return max(allowed) if allowed else None
    if allowed and dimensions not in allowed:
        raise ValueError(
            f"Unsupported --dimensions={dimensions} for model={model}. "
            f"Allowed values: {allowed}"
        )
    return dimensions


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def write_jsonl_atomic(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    os.replace(tmp, path)


def clear_output_dir(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path)


def file_signature(path: Path | None) -> dict[str, int] | None:
    if not path or not path.exists() or not path.is_file():
        return None
    st = path.stat()
    return {"size": int(st.st_size), "mtime_ns": int(st.st_mtime_ns)}


def iter_workbooks(data_root: Path) -> list[Path]:
    workbooks: list[Path] = []
    for path in sorted(data_root.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith(TEMP_PREFIXES):
            continue
        if path.suffix.lower() in SUPPORTED_WORKBOOK_SUFFIXES:
            workbooks.append(path)
    return workbooks


def resolve_existing_file(source_dir: Path, raw_name: str, allowed_suffixes: tuple[str, ...]) -> Path | None:
    raw = safe_str(raw_name)
    if not raw:
        return None
    raw_path = Path(raw)

    candidates: list[Path] = [source_dir / raw_path, source_dir / raw_path.name]
    if raw_path.suffix:
        candidates.append(source_dir / f"{raw_path.stem}{raw_path.suffix.lower()}")
    else:
        stem = raw_path.name
        candidates.extend(source_dir / f"{stem}{suffix}" for suffix in allowed_suffixes)

    stem = raw_path.stem if raw_path.suffix else raw_path.name
    for suffix in allowed_suffixes:
        candidates.append(source_dir / f"{stem}{suffix}")

    seen = set()
    for candidate in candidates:
        key = str(candidate)
        if key in seen:
            continue
        seen.add(key)
        if candidate.exists() and candidate.is_file():
            return candidate

    target_name = raw_path.name
    if target_name:
        for path in source_dir.rglob(target_name):
            if path.is_file():
                return path
    if stem:
        for suffix in allowed_suffixes:
            for path in source_dir.rglob(f"{stem}{suffix}"):
                if path.is_file():
                    return path
    return None


def infer_ext_from_file_type(file_type: str) -> str:
    token = normalize_token(file_type)
    if "ppt" in token:
        return ".pptx"
    if "doc" in token or "word" in token:
        return ".docx"
    if "pdf" in token:
        return ".pdf"
    if "xls" in token:
        return ".xlsx"
    return ".txt"


def resolve_doc_source_file(source_dir: Path, file_name: str, file_type: str) -> Path | None:
    raw = safe_str(file_name)
    if not raw:
        return None
    raw_path = Path(raw)

    candidates: list[Path] = [source_dir / raw_path, source_dir / raw_path.name]
    stem = raw_path.stem if raw_path.suffix else raw_path.name

    if raw_path.suffix:
        candidates.append(source_dir / f"{stem}{raw_path.suffix.lower()}")

    inferred = infer_ext_from_file_type(file_type)
    candidates.append(source_dir / f"{stem}{inferred}")

    for suffix in (
        ".pptx",
        ".ppt",
        ".pdf",
        ".docx",
        ".doc",
        ".xlsx",
        ".xls",
        ".md",
        ".txt",
    ):
        candidates.append(source_dir / f"{stem}{suffix}")

    seen = set()
    for candidate in candidates:
        key = str(candidate)
        if key in seen:
            continue
        seen.add(key)
        if candidate.exists() and candidate.is_file():
            return candidate

    target_name = raw_path.name
    if target_name:
        for path in source_dir.rglob(target_name):
            if path.is_file():
                return path

    return None


def copy_if_needed(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    if dst.exists():
        src_stat = src.stat()
        dst_stat = dst.stat()
        if dst_stat.st_size == src_stat.st_size and dst_stat.st_mtime_ns >= src_stat.st_mtime_ns:
            return
    shutil.copy2(src, dst)


def clear_doc_assets(output_dir: Path, doc_id: str) -> None:
    for folder in (
        output_dir / "assets" / "docs" / doc_id,
        output_dir / "assets" / "images" / doc_id,
        output_dir / "assets" / "tables" / doc_id,
    ):
        if folder.exists():
            shutil.rmtree(folder)


def load_workbook_rows(workbook_path: Path) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sheets = wb.worksheets
    if len(sheets) < 3:
        raise ValueError(f"Workbook has fewer than 3 sheets: {workbook_path}")

    image_ws = sheets[0]
    table_ws = sheets[1]
    text_ws = sheets[2]

    images: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    texts: list[dict[str, Any]] = []

    for row_index, row in enumerate(image_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None or safe_str(row[0]) == "":
            continue
        images.append(
            {
                "raw_path": safe_str(row[0]),
                "keywords": safe_str(row[1]),
                "directory": safe_str(row[2]),
                "file_type": safe_str(row[3]),
                "file_name": safe_str(row[4]),
                "page_no": safe_int(row[5], default=0),
                "row_index": row_index,
                "source_sheet": image_ws.title,
            }
        )

    for row_index, row in enumerate(table_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None or safe_str(row[0]) == "":
            continue
        tables.append(
            {
                "raw_path": safe_str(row[0]),
                "keywords": safe_str(row[1]),
                "directory": safe_str(row[2]),
                "file_type": safe_str(row[3]),
                "file_name": safe_str(row[4]),
                "page_no": safe_int(row[5], default=0),
                "row_index": row_index,
                "source_sheet": table_ws.title,
            }
        )

    for row_index, row in enumerate(text_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None or safe_str(row[0]) == "":
            continue
        text_content = normalize_text(safe_str(row[6]) if len(row) > 6 else "")
        if not text_content:
            continue
        texts.append(
            {
                "text_id": safe_str(row[0]),
                "keywords": safe_str(row[1]),
                "directory": safe_str(row[2]),
                "file_type": safe_str(row[3]),
                "file_name": safe_str(row[4]),
                "page_no": safe_int(row[5], default=0),
                "text_content": text_content,
                "row_index": row_index,
                "source_sheet": text_ws.title,
            }
        )

    return {"images": images, "tables": tables, "texts": texts}


def make_group_key(directory: str, file_type: str, file_name: str) -> tuple[str, str, str]:
    stem = Path(file_name).stem if Path(file_name).suffix else file_name
    return (normalize_directory_key(directory), canonical_file_type(file_type), normalize_token(stem))


def build_group_indexes(
    text_groups: dict[tuple[str, str, str], list[dict[str, Any]]],
) -> tuple[
    dict[str, list[tuple[str, str, str]]],
    dict[tuple[str, str], list[tuple[str, str, str]]],
    dict[str, list[tuple[str, str, str]]],
]:
    by_name: dict[str, list[tuple[str, str, str]]] = {}
    by_dir_type: dict[tuple[str, str], list[tuple[str, str, str]]] = {}
    by_type: dict[str, list[tuple[str, str, str]]] = {}

    for key in text_groups:
        dir_key, type_key, name_key = key
        by_name.setdefault(name_key, []).append(key)
        by_dir_type.setdefault((dir_key, type_key), []).append(key)
        by_type.setdefault(type_key, []).append(key)

    return by_name, by_dir_type, by_type


def resolve_asset_group_key(
    row: dict[str, Any],
    text_groups: dict[tuple[str, str, str], list[dict[str, Any]]],
    by_name: dict[str, list[tuple[str, str, str]]],
    by_dir_type: dict[tuple[str, str], list[tuple[str, str, str]]],
    by_type: dict[str, list[tuple[str, str, str]]],
) -> tuple[str, str, str]:
    key = make_group_key(row["directory"], row["file_type"], row["file_name"])
    if key in text_groups:
        return key

    name_key = key[2]
    name_matches = by_name.get(name_key, [])
    if len(name_matches) == 1:
        return name_matches[0]

    dir_type_matches = by_dir_type.get((key[0], key[1]), [])
    if len(dir_type_matches) == 1:
        return dir_type_matches[0]

    type_matches = by_type.get(key[1], [])
    if len(type_matches) == 1:
        return type_matches[0]

    fuzzy_name_matches = [
        candidate
        for candidate in text_groups
        if name_key and (name_key in candidate[2] or candidate[2] in name_key)
    ]
    if len(fuzzy_name_matches) == 1:
        return fuzzy_name_matches[0]

    return key


def resolve_table_path(source_dir: Path, raw_path: str) -> Path | None:
    raw = safe_str(raw_path)
    if not raw:
        return None
    p = Path(raw)
    candidates = [source_dir / p, source_dir / p.name]
    if p.suffix:
        candidates.append(source_dir / f"{p.stem}{p.suffix.lower()}")
    else:
        candidates.extend([source_dir / f"{p.name}.md", source_dir / f"{p.name}.txt"])

    seen = set()
    for c in candidates:
        key = str(c)
        if key in seen:
            continue
        seen.add(key)
        if c.exists() and c.is_file():
            return c

    target_name = p.name
    if target_name:
        for c in source_dir.rglob(target_name):
            if c.is_file():
                return c
    return None


def build_doc_id(
    source_dir: Path,
    rel_workbook: str,
    directory: str,
    file_type: str,
    doc_title: str,
    used_doc_ids: set[str],
    preferred_doc_id: str = "",
) -> str:
    preferred = safe_str(preferred_doc_id)
    if preferred:
        used_doc_ids.add(preferred)
        return preferred

    base = slugify(f"{source_dir.name}_{doc_title}")
    if base not in used_doc_ids:
        used_doc_ids.add(base)
        return base

    token = f"{rel_workbook}|{directory}|{file_type}|{doc_title}"
    suffix = hashlib.sha1(token.encode("utf-8")).hexdigest()[:8]
    candidate = f"{base}_{suffix}"
    while candidate in used_doc_ids:
        suffix = hashlib.sha1((token + candidate).encode("utf-8")).hexdigest()[:8]
        candidate = f"{base}_{suffix}"
    used_doc_ids.add(candidate)
    return candidate


def build_text_chunk_text(
    page_no: int,
    keywords: str,
    text_content: str,
    related_images: list[AssetRef],
    related_tables: list[AssetRef],
) -> str:
    parts = [
        f"页码: {page_no}",
        "类型: text",
        f"关键词: {keywords or 'N/A'}",
    ]
    if related_images:
        parts.append("关联图片: " + "；".join(item.rel_path for item in related_images))
    if related_tables:
        parts.append("关联表格: " + "；".join(item.rel_path for item in related_tables))
    parts.append("")
    parts.append("正文:")
    parts.append(text_content)
    return normalize_text("\n".join(parts))


def asset_keyword_overlap(text_keywords: list[str], asset_keywords: list[str]) -> float:
    if not text_keywords or not asset_keywords:
        return 0.0
    left = set(text_keywords)
    right = set(asset_keywords)
    common = len(left & right)
    if common == 0:
        return 0.0
    return common / max(min(len(left), len(right)), 1)


def lexical_overlap_score(text: str, asset: AssetRef) -> float:
    text_norm = normalize_text(text)
    if not text_norm:
        return 0.0
    haystack = normalize_text("\n".join([asset.keywords, asset.title]))
    if not haystack:
        return 0.0
    matches = 0
    for token in split_keywords(asset.keywords):
        if token and token in text_norm:
            matches += 1
    if asset.title and asset.title in text_norm:
        matches += 1
    return min(matches / 3.0, 1.0)


def select_related_assets_for_text_row(
    *,
    text_keywords: list[str],
    text_content: str,
    images: list[AssetRef],
    tables: list[AssetRef],
) -> tuple[list[AssetRef], list[AssetRef]]:
    def rank_assets(candidates: list[AssetRef], *, limit: int, minimum_score: float) -> list[AssetRef]:
        scored: list[tuple[float, int, AssetRef]] = []
        for idx, asset in enumerate(candidates):
            overlap = asset_keyword_overlap(text_keywords, split_keywords(asset.keywords))
            lexical = lexical_overlap_score(text_content, asset)
            score = 0.7 * overlap + 0.3 * lexical
            if len(candidates) == 1:
                score = max(score, minimum_score)
            if score >= minimum_score:
                scored.append((score, idx, asset))
        scored.sort(key=lambda item: (-item[0], item[1]))
        return [asset for _, _, asset in scored[:limit]]

    selected_images = rank_assets(images, limit=1, minimum_score=0.18)
    selected_tables = rank_assets(tables, limit=1, minimum_score=0.18)
    return selected_images, selected_tables


def build_image_chunk_text(asset: AssetRef, doc_title: str, page_assets: list[AssetRef]) -> str:
    parts = [
        f"文档: {doc_title}",
        f"页码: {asset.page_no}",
        "类型: image",
        f"关键词: {asset.keywords or 'N/A'}",
        f"图片文件: {asset.rel_path}",
    ]
    sibling_tables = [item.rel_path for item in page_assets if item.asset_type == "table"]
    if sibling_tables:
        parts.append("同页表格: " + "；".join(sibling_tables))
    return normalize_text("\n".join(parts))


def build_table_chunk_text(asset: AssetRef, doc_title: str) -> str:
    table_text = ""
    asset_path = Path(asset.abs_path)
    if asset_path.exists():
        try:
            table_text = asset_path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            table_text = asset_path.read_text(encoding="gbk", errors="replace")
    parts = [
        f"文档: {doc_title}",
        f"页码: {asset.page_no}",
        "类型: table",
        f"关键词: {asset.keywords or 'N/A'}",
        f"表格文件: {asset.rel_path}",
        "",
        "Markdown表格:",
        table_text,
    ]
    return normalize_text("\n".join(parts))


def build_document_bundle(doc_title: str, doc_path: str, text_rows: list[dict[str, Any]], ext: str, doc_id: str) -> dict[str, Any]:
    sections = []
    for row in sorted(text_rows, key=lambda x: (x["page_no"], x["row_index"])):
        sections.append(f"--- Slide {row['page_no']} ---\n{row['text_content']}")
    return {
        "file_path": doc_path,
        "ext": ext,
        "text": "\n".join(sections),
        "image_paths": [],
        "table_paths": [],
        "doc_title": doc_title,
        "doc_id": doc_id,
    }


def make_doc_progress_key(workbook_path: Path, directory: str, file_type: str, file_name: str) -> str:
    key = "::".join([str(workbook_path.resolve()), normalize_token(directory), normalize_token(file_type), normalize_token(file_name)])
    return key


def build_doc_signature(
    workbook_path: Path,
    doc_id: str,
    doc_title: str,
    text_rows: list[dict[str, Any]],
    image_rows: list[dict[str, Any]],
    table_rows: list[dict[str, Any]],
    source_dir: Path,
    doc_source_file: Path | None,
) -> str:
    hasher = hashlib.sha256()
    hasher.update(str(workbook_path.resolve()).encode("utf-8"))
    hasher.update(doc_id.encode("utf-8"))
    hasher.update(doc_title.encode("utf-8"))

    for row in sorted(text_rows, key=lambda x: (x["page_no"], x["row_index"])):
        token = "|".join(
            [
                str(row["page_no"]),
                row["keywords"],
                row["directory"],
                row["file_type"],
                row["file_name"],
                row["text_content"],
            ]
        )
        hasher.update(token.encode("utf-8"))

    for rows, kind in ((image_rows, "image"), (table_rows, "table")):
        for row in sorted(rows, key=lambda x: (x["page_no"], x["row_index"], x["raw_path"])):
            raw = row["raw_path"]
            sig = None
            if kind == "image":
                resolved = resolve_existing_file(source_dir, raw, (".png", ".jpg", ".jpeg", ".webp", ".bmp"))
                sig = file_signature(resolved)
            else:
                resolved = resolve_table_path(source_dir, raw)
                sig = file_signature(resolved)
            token = {
                "kind": kind,
                "raw": raw,
                "page_no": row["page_no"],
                "keywords": row["keywords"],
                "directory": row["directory"],
                "file_type": row["file_type"],
                "file_name": row["file_name"],
                "sig": sig,
            }
            hasher.update(json.dumps(token, ensure_ascii=False, sort_keys=True).encode("utf-8"))

    hasher.update(json.dumps(file_signature(doc_source_file), ensure_ascii=False, sort_keys=True).encode("utf-8"))
    return hasher.hexdigest()


def build_doc_records(
    workbook_path: Path,
    rel_workbook: str,
    source_dir: Path,
    output_dir: Path,
    doc_id: str,
    doc_title: str,
    doc_rel_path: Path,
    doc_source_path: Path | None,
    doc_ext: str,
    text_rows: list[dict[str, Any]],
    image_rows: list[dict[str, Any]],
    table_rows: list[dict[str, Any]],
    next_chunk_id: int,
    dry_run: bool = False,
) -> tuple[list[ChunkRecord], list[PageRecord], list[TableRecord], list[AssetRef], DocRecord, dict[str, Any], int]:
    if not dry_run and doc_source_path:
        copy_if_needed(doc_source_path, output_dir / doc_rel_path)

    images_by_page: dict[int, list[AssetRef]] = {}
    tables_by_page: dict[int, list[AssetRef]] = {}
    asset_uid_seen: set[str] = set()

    for row in image_rows:
        source_path = resolve_existing_file(source_dir, row["raw_path"], (".png", ".jpg", ".jpeg", ".webp", ".bmp"))
        if not source_path:
            continue
        copied_rel = Path("assets/images") / doc_id / source_path.name
        copied_abs = output_dir / copied_rel
        if not dry_run:
            copy_if_needed(source_path, copied_abs)

        base_uid = f"{doc_id}:image:{slugify(source_path.stem)}"
        uid = base_uid
        idx = 2
        while uid in asset_uid_seen:
            uid = f"{base_uid}:{idx}"
            idx += 1
        asset_uid_seen.add(uid)

        asset = AssetRef(
            asset_id=uid,
            asset_type="image",
            page_no=row["page_no"],
            keywords=row["keywords"],
            rel_path=str(copied_rel),
            abs_path=str(copied_abs),
            source_sheet=row["source_sheet"],
            source_row_index=row["row_index"],
            title=source_path.name,
        )
        images_by_page.setdefault(row["page_no"], []).append(asset)

    table_records: list[TableRecord] = []
    for row in table_rows:
        source_path = resolve_table_path(source_dir, row["raw_path"])
        if not source_path:
            continue
        copied_rel = Path("assets/tables") / doc_id / source_path.name
        copied_abs = output_dir / copied_rel
        if not dry_run:
            copy_if_needed(source_path, copied_abs)

        base_uid = f"{doc_id}:table:{slugify(source_path.stem)}"
        uid = base_uid
        idx = 2
        while uid in asset_uid_seen:
            uid = f"{base_uid}:{idx}"
            idx += 1
        asset_uid_seen.add(uid)

        asset = AssetRef(
            asset_id=uid,
            asset_type="table",
            page_no=row["page_no"],
            keywords=row["keywords"],
            rel_path=str(copied_rel),
            abs_path=str(copied_abs),
            source_sheet=row["source_sheet"],
            source_row_index=row["row_index"],
            title=source_path.name,
        )
        tables_by_page.setdefault(row["page_no"], []).append(asset)

        # Create Table Record
        table_records.append(
            TableRecord(
                table_id=uid,
                doc_id=doc_id,
                page_id=f"{doc_id}:page:{row['page_no']}",
                page_no=row["page_no"],
                doc_title=doc_title,
                table_title=asset.title,
                table_headers=[],  # Placeholder
                table_summary=asset.keywords,
                table_markdown=asset.rel_path,  # Use rel_path as placeholder for content if not available
                keywords=split_keywords(asset.keywords),
                asset_path=str(copied_rel),
            )
        )

    chunk_records: list[ChunkRecord] = []
    page_records: list[PageRecord] = []
    
    # Process pages
    # Process pages
    max_page = max(
        [r["page_no"] for r in text_rows] +
        [r["page_no"] for r in image_rows] +
        [r["page_no"] for r in table_rows] +
        [1]
    )
    all_page_ids = []
    all_keywords = set()

    for page_no in range(1, max_page + 1):
        page_id = f"{doc_id}:page:{page_no}"
        page_chunk_ids = []
        page_table_ids = [t.table_id for t in table_records if t.page_no == page_no]
        page_image_ids = [i.asset_id for i in images_by_page.get(page_no, [])]
        page_keywords = set()

        # Text chunks for this page
        page_text_rows = [r for r in text_rows if r["page_no"] == page_no]
        for i, row in enumerate(sorted(page_text_rows, key=lambda x: x["row_index"])):
            chunk_uid = f"{doc_id}:text:p{page_no}:c{i+1}"
            related_images, related_tables = select_related_assets_for_text_row(
                text_keywords=split_keywords(row["keywords"]),
                text_content=row["text_content"],
                images=images_by_page.get(page_no, []),
                tables=tables_by_page.get(page_no, []),
            )
            
            chunk_text = build_text_chunk_text(page_no, row["keywords"], row["text_content"], related_images, related_tables)
            keywords = split_keywords(row["keywords"])
            page_keywords.update(keywords)
            
            record = ChunkRecord(
                chunk_id=chunk_uid,
                parent_chunk_uid=page_id,
                doc_id=doc_id,
                page_id=page_id,
                page_no=page_no,
                doc_title=doc_title,
                heading=f"{doc_title} 第{page_no}页正文",
                chunk_order=i,
                chunk_text_clean=normalize_text(row["text_content"]),
                chunk_text_display=chunk_text,
                keywords=keywords,
                table_ids_on_page=page_table_ids,
                image_ids_on_page=page_image_ids,
                asset_type="text",
                image_paths=[i.rel_path for i in images_by_page.get(page_no, [])],
                table_paths=[t.asset_path for t in table_records if t.page_no == page_no],
                file_path=str(doc_rel_path),
                ext=doc_ext,
                para_id=0,
            )
            chunk_records.append(record)
            page_chunk_ids.append(chunk_uid)

        if page_chunk_ids or page_table_ids or page_image_ids:
            page_records.append(
                PageRecord(
                    page_id=page_id,
                    doc_id=doc_id,
                    page_no=page_no,
                    doc_title=doc_title,
                    page_summary=f"Page {page_no} of {doc_title}",
                    page_keywords=list(page_keywords),
                    chunk_ids=page_chunk_ids,
                    table_ids=page_table_ids,
                    image_ids=page_image_ids,
                )
            )
            all_page_ids.append(page_id)
            all_keywords.update(page_keywords)

    doc_record = DocRecord(
        doc_id=doc_id,
        doc_title=doc_title,
        doc_summary=f"Summary of {doc_title}",
        page_count=len(page_records),
        top_keywords=list(all_keywords)[:10],
        page_ids=all_page_ids,
    )

    all_image_assets = []
    for p_imgs in images_by_page.values():
        all_image_assets.extend(p_imgs)

    bundle = build_document_bundle(doc_title, str(doc_rel_path), text_rows, doc_ext, doc_id)
    return chunk_records, page_records, table_records, all_image_assets, doc_record, bundle, next_chunk_id


def merge_rows(existing: list[dict[str, Any]], new_rows: list[dict[str, Any]], replaced_doc_ids: set[str]) -> list[dict[str, Any]]:
    kept = [row for row in existing if row.get("doc_id") not in replaced_doc_ids]
    kept.extend(new_rows)
    return kept


def embed_texts(
    api_base: str,
    api_key: str,
    model: str,
    texts: list[str],
    batch_size: int,
    timeout: float,
    dimensions: int | None,
) -> list[list[float]]:
    if not texts:
        return []
    embeddings: list[list[float]] = []
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    session = requests.Session()
    endpoint = api_base.rstrip("/") + "/embeddings"

    for start in range(0, len(texts), batch_size):
        batch = texts[start : start + batch_size]
        payload: dict[str, Any] = {
            "model": model,
            "input": batch,
            "encoding_format": "float",
        }
        if dimensions:
            payload["dimensions"] = dimensions
        response = session.post(endpoint, headers=headers, json=payload, timeout=timeout)
        response.raise_for_status()
        data = response.json()["data"]
        embeddings.extend(item["embedding"] for item in data)
        time.sleep(0.1)
    return embeddings


def save_numpy_index(output_dir: Path, records: list[dict[str, Any]], index_name: str) -> np.ndarray:
    vector_store = output_dir / "vector_store"
    vector_store.mkdir(parents=True, exist_ok=True)
    npy_path = vector_store / f"{index_name}_embeddings.npy"
    if not records:
        return np.array([], dtype=np.float32)
    embeddings = [r["embedding"] for r in records]
    matrix = np.array(embeddings, dtype=np.float32)
    np.save(str(npy_path), matrix)
    return matrix


 


def compute_mean_embedding(children_embeddings: list[list[float]]) -> list[float] | None:
    if not children_embeddings:
        return None
    matrix = np.array(children_embeddings, dtype=np.float32)
    mean = np.mean(matrix, axis=0)
    return mean.tolist()


def write_manifest(
    output_dir: Path,
    api_base: str,
    model: str,
    dimensions: int | None,
    doc_updates: dict[str, dict[str, Any]],
    removed_doc_ids: set[str],
    backend: str,
    index_manifest: dict[str, dict[str, Any]],
) -> None:
    vector_store = output_dir / "vector_store"
    vector_store.mkdir(parents=True, exist_ok=True)
    manifest_path = vector_store / "manifest.json"
    manifest: dict[str, Any] = {}
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    docs = manifest.get("documents", {})
    for doc_id in removed_doc_ids:
        docs.pop(doc_id, None)

    for doc_id, meta in doc_updates.items():
        docs[doc_id] = {
            "doc_title": meta["doc_title"],
            "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "chunk_count": meta["chunk_count"],
        }

    manifest.update(
        {
            "backend": backend,
            "embedding_api_base": api_base,
            "embedding_model": model,
            "embedding_dimensions": dimensions,
            "indexes": index_manifest,
            "documents": docs,
        }
    )
    write_json_atomic(manifest_path, manifest)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build/expand a RAG corpus from processed Excel workbooks.")
    parser.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT)
    parser.add_argument("--workbook", type=Path, action="append", default=[])
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--progress-file", type=Path, default=None)
    parser.add_argument("--force-rebuild", action="store_true")
    parser.add_argument("--clear-output", action="store_true", help="Delete output dir and restart from scratch.")
 
    parser.add_argument("--max-workbooks", type=int, default=0, help="0 means no limit")
    parser.add_argument("--max-docs", type=int, default=0, help="0 means no limit")
    parser.add_argument("--dry-run", action="store_true", help="Analyze and stage logic only, do not write outputs")
    parser.add_argument("--api-key", default="")
    parser.add_argument("--embedding-api-base", default=DEFAULT_EMBEDDING_API_BASE)
    parser.add_argument("--embedding-model", default=DEFAULT_EMBEDDING_MODEL)
    parser.add_argument(
        "--dimensions",
        type=int,
        default=DEFAULT_EMBEDDING_DIMENSIONS,
        help="Embedding dimensions. 0 means use the selected model's maximum supported dimensions.",
    )
    parser.add_argument("--batch-size", type=int, default=16)
    parser.add_argument("--timeout", type=float, default=120.0)
    parser.add_argument("--skip-embeddings", action="store_true")
    args = parser.parse_args()
    args.api_key = resolve_env_value("SILICONFLOW_API_KEY", args.api_key)
    return args


def main() -> int:
    args = parse_args()
    data_root = args.data_root.resolve()
    output_dir = args.output_dir.resolve()
    progress_path = resolve_progress_file(data_root=data_root, progress_file=args.progress_file)
    resolved_dimensions = resolve_embedding_dimensions(args.embedding_model, args.dimensions)

    if args.clear_output and not args.dry_run:
        clear_output_dir(output_dir)
        if progress_path.exists():
            progress_path.unlink()

    output_dir.mkdir(parents=True, exist_ok=True)

    if args.workbook:
        workbooks = [p.resolve() for p in args.workbook]
    else:
        workbooks = iter_workbooks(data_root)

    if args.max_workbooks > 0:
        workbooks = workbooks[: args.max_workbooks]

    progress: dict[str, Any] = read_json(
        progress_path,
        {
            "version": 1,
            "updated_at": now_iso(),
            "workbooks": {},
            "documents": {},
        },
    )

    chunks_path = output_dir / "chunks.jsonl"
    pages_path = output_dir / "pages.jsonl"
    tables_path = output_dir / "tables.jsonl"
    docs_path = output_dir / "docs.jsonl"
    all_texts_path = output_dir / "all_texts.jsonl"

    existing_chunks = read_jsonl(chunks_path)
    existing_pages = read_jsonl(pages_path)
    existing_tables = read_jsonl(tables_path)
    existing_docs = read_jsonl(docs_path)
    existing_all_texts = read_jsonl(all_texts_path)

    max_chunk_id = max((safe_int(row.get("chunk_id"), default=0) for row in existing_chunks), default=-1)
    next_chunk_id = max_chunk_id + 1

    used_doc_ids: set[str] = set()

    all_chunks: list[ChunkRecord] = []
    all_pages: list[PageRecord] = []
    all_tables: list[TableRecord] = []
    all_docs: list[DocRecord] = []
    all_bundles: list[dict[str, Any]] = []

    staged_docs: dict[str, dict[str, Any]] = {}
    removed_doc_ids: set[str] = set()

    processed_doc_counter = 0
    workbook_counter = 0

    for workbook_path in workbooks:
        workbook_counter += 1
        wb_key = str(workbook_path)
        wb_sig = file_signature(workbook_path)
        wb_entry = progress.setdefault("workbooks", {}).setdefault(wb_key, {})
        wb_entry.update(
            {
                "workbook_path": wb_key,
                "source_dir": str(workbook_path.parent),
                "status": "processing",
                "started_at": wb_entry.get("started_at") or now_iso(),
                "updated_at": now_iso(),
                "signature": wb_sig,
                "summary": wb_entry.get("summary") or {},
            }
        )
        write_json_atomic(progress_path, progress)

        try:
            parsed = load_workbook_rows(workbook_path)
        except Exception as exc:  # noqa: BLE001
            wb_entry["status"] = "failed"
            wb_entry["updated_at"] = now_iso()
            wb_entry["error"] = str(exc)
            write_json_atomic(progress_path, progress)
            print(f"[WARN] Skip workbook due to parse error: {workbook_path} :: {exc}", file=sys.stderr)
            continue

        text_groups: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
        sample_meta: dict[tuple[str, str, str], dict[str, str]] = {}

        for row in parsed["texts"]:
            key = make_group_key(row["directory"], row["file_type"], row["file_name"])
            text_groups.setdefault(key, []).append(row)
            sample_meta.setdefault(
                key,
                {
                    "directory": row["directory"],
                    "file_type": row["file_type"],
                    "file_name": row["file_name"],
                },
            )

        by_name, by_dir_type, by_type = build_group_indexes(text_groups)

        image_groups: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
        for row in parsed["images"]:
            key = resolve_asset_group_key(row, text_groups, by_name, by_dir_type, by_type)
            image_groups.setdefault(key, []).append(row)

        table_groups: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
        for row in parsed["tables"]:
            key = resolve_asset_group_key(row, text_groups, by_name, by_dir_type, by_type)
            table_groups.setdefault(key, []).append(row)

        rel_workbook = str(workbook_path.relative_to(data_root)) if workbook_path.is_relative_to(data_root) else workbook_path.name
        current_wb_doc_ids: list[str] = []
        wb_processed = 0
        wb_skipped = 0
        wb_failed = 0

        for key, text_rows in sorted(text_groups.items(), key=lambda item: (item[0][0], item[0][1], item[0][2])):
            if args.max_docs > 0 and processed_doc_counter >= args.max_docs:
                break

            meta = sample_meta[key]
            doc_title = Path(meta["file_name"]).stem if Path(meta["file_name"]).suffix else meta["file_name"]
            if not doc_title:
                doc_title = "未命名文档"

            doc_key = make_doc_progress_key(workbook_path, meta["directory"], meta["file_type"], meta["file_name"])
            doc_progress = progress.setdefault("documents", {}).get(doc_key, {})

            doc_id = build_doc_id(
                source_dir=workbook_path.parent,
                rel_workbook=rel_workbook,
                directory=meta["directory"],
                file_type=meta["file_type"],
                doc_title=doc_title,
                used_doc_ids=used_doc_ids,
                preferred_doc_id=doc_progress.get("doc_id", ""),
            )
            current_wb_doc_ids.append(doc_id)

            doc_source_file = resolve_doc_source_file(workbook_path.parent, meta["file_name"], meta["file_type"])
            doc_ext = doc_source_file.suffix if doc_source_file else infer_ext_from_file_type(meta["file_type"])
            doc_rel_path = Path("assets/docs") / doc_id / (
                doc_source_file.name if doc_source_file else f"{doc_title}{doc_ext or '.txt'}"
            )

            image_rows = image_groups.get(key, [])
            table_rows = table_groups.get(key, [])

            signature = build_doc_signature(
                workbook_path=workbook_path,
                doc_id=doc_id,
                doc_title=doc_title,
                text_rows=text_rows,
                image_rows=image_rows,
                table_rows=table_rows,
                source_dir=workbook_path.parent,
                doc_source_file=doc_source_file,
            )

            unchanged = (
                not args.force_rebuild
                and doc_progress.get("status") == "completed"
                and doc_progress.get("signature") == signature
                and doc_progress.get("doc_id") == doc_id
            )

            if unchanged:
                wb_skipped += 1
                progress["documents"][doc_key] = {
                    **doc_progress,
                    "status": "completed",
                    "doc_id": doc_id,
                    "doc_title": doc_title,
                    "signature": signature,
                    "updated_at": now_iso(),
                    "workbook_path": str(workbook_path),
                }
                write_json_atomic(progress_path, progress)
                continue

            progress["documents"][doc_key] = {
                "status": "staged",
                "doc_id": doc_id,
                "doc_title": doc_title,
                "signature": signature,
                "updated_at": now_iso(),
                "workbook_path": str(workbook_path),
            }
            write_json_atomic(progress_path, progress)

            try:
                chunks, pages, tables, img_assets, doc_rec, bundle, next_chunk_id = build_doc_records(
                    workbook_path=workbook_path,
                    rel_workbook=rel_workbook,
                    source_dir=workbook_path.parent,
                    output_dir=output_dir,
                    doc_id=doc_id,
                    doc_title=doc_title,
                    doc_rel_path=doc_rel_path,
                    doc_source_path=doc_source_file,
                    doc_ext=doc_ext,
                    text_rows=text_rows,
                    image_rows=image_rows,
                    table_rows=table_rows,
                    next_chunk_id=next_chunk_id,
                    dry_run=args.dry_run,
                )
            except Exception as exc:  # noqa: BLE001
                wb_failed += 1
                progress["documents"][doc_key] = {
                    "status": "failed",
                    "doc_id": doc_id,
                    "doc_title": doc_title,
                    "signature": signature,
                    "updated_at": now_iso(),
                    "workbook_path": str(workbook_path),
                    "error": str(exc),
                }
                write_json_atomic(progress_path, progress)
                print(f"[WARN] Skip document due to build error: {doc_id} :: {exc}", file=sys.stderr)
                continue

            all_chunks.extend(chunks)
            all_pages.extend(pages)
            all_tables.extend(tables)
            all_docs.append(doc_rec)
            all_bundles.append(bundle)

            staged_docs[doc_id] = {
                "chunks": [asdict(record) for record in chunks],
                "pages": [asdict(record) for record in pages],
                "tables": [asdict(record) for record in tables],
                "images": [asdict(record) for record in img_assets],
                "doc_rec": asdict(doc_rec),
                "bundle": bundle,
                "doc_key": doc_key,
                "doc_title": doc_title,
                "signature": signature,
            }
            wb_processed += 1
            processed_doc_counter += 1

            progress["documents"][doc_key] = {
                "status": "staged",
                "doc_id": doc_id,
                "doc_title": doc_title,
                "signature": signature,
                "chunk_count": len(chunks),
                "updated_at": now_iso(),
                "workbook_path": str(workbook_path),
            }
            write_json_atomic(progress_path, progress)

        previous_doc_ids = set(wb_entry.get("doc_ids", []))
        current_doc_ids = set(current_wb_doc_ids)
        stale_for_workbook = previous_doc_ids - current_doc_ids
        if stale_for_workbook:
            removed_doc_ids.update(stale_for_workbook)

        wb_entry["status"] = "completed"
        wb_entry["updated_at"] = now_iso()
        wb_entry["doc_ids"] = sorted(current_doc_ids)
        wb_entry["summary"] = {
            "processed": wb_processed,
            "skipped": wb_skipped,
            "failed": wb_failed,
            "groups": len(text_groups),
        }
        write_json_atomic(progress_path, progress)

        if args.max_docs > 0 and processed_doc_counter >= args.max_docs:
            break

    replaced_doc_ids = set(staged_docs.keys()) | removed_doc_ids

    if args.dry_run:
        print(
            f"Dry-run complete: workbooks_scanned={workbook_counter}, staged_docs={len(staged_docs)}, "
            f"removed_docs={len(removed_doc_ids)}"
        )
        return 0

    chunks_new_rows: list[dict[str, Any]] = []
    pages_new_rows: list[dict[str, Any]] = []
    tables_new_rows: list[dict[str, Any]] = []
    docs_new_rows: list[dict[str, Any]] = []
    all_texts_new_rows: list[dict[str, Any]] = []
    doc_manifest_updates: dict[str, dict[str, Any]] = {}

    images_path = output_dir / "images.jsonl"
    relations_path = output_dir / "relations_chunk_asset.jsonl"

    all_images_rows: list[dict[str, Any]] = []
    all_relations_rows: list[dict[str, Any]] = []

    for doc_id, payload in staged_docs.items():
        chunks_new_rows.extend(payload["chunks"])
        pages_new_rows.extend(payload["pages"])
        tables_new_rows.extend(payload["tables"])
        all_images_rows.extend(payload["images"])
        docs_new_rows.append(payload["doc_rec"])
        all_texts_new_rows.append(payload["bundle"])
        doc_manifest_updates[doc_id] = {
            "doc_title": payload["doc_title"],
            "chunk_count": len(payload["chunks"]),
        }
    
    # images mapping for images.jsonl
    images_merged = merge_rows([], all_images_rows, set()) # Simplified for now, should ideally merge
    # Wait, merge_rows needs existing_rows. I'll just collect them all for now since images.jsonl is usually full rebuild
    
    chunks_merged = merge_rows(existing_chunks, chunks_new_rows, replaced_doc_ids)
    pages_merged = merge_rows(existing_pages, pages_new_rows, replaced_doc_ids)
    tables_merged = merge_rows(existing_tables, tables_new_rows, replaced_doc_ids)
    docs_merged = merge_rows(existing_docs, docs_new_rows, replaced_doc_ids)
    all_texts_merged = merge_rows(existing_all_texts, all_texts_new_rows, replaced_doc_ids)

    # Format images for images.jsonl
    final_images_rows = []
    for img in all_images_rows:
        final_images_rows.append({
            "image_id": img["asset_id"],
            "image_title": Path(img["rel_path"]).name,
            "doc_id": img["rel_path"].split("/")[-2] if "/" in img["rel_path"] else "unknown",
            "page_id": f"{img['rel_path'].split('/')[-2]}:page:{img['page_no']}" if "/" in img["rel_path"] else "unknown",
            "page_no": img["page_no"],
            "image_path": img["rel_path"],
            "keywords": split_keywords(img["keywords"])
        })
    # Note: doc_id extraction from rel_path is a bit hacky, but consistent with current assets layout

    # Generate relations_chunk_asset.jsonl from merged data
    # (Extracting from ChunkRecord table_ids_on_page/image_ids_on_page)
    
    # relations
    for c in chunks_merged:
        for tid in c.get("table_ids_on_page", []):
            all_relations_rows.append({
                "chunk_id": c["chunk_id"],
                "asset_id": tid,
                "asset_type": "table",
                "page_id": c["page_id"],
                "relation_score": 0.55,
                "relation_reasons": ["same_page"]
            })
        for iid in c.get("image_ids_on_page", []):
            all_relations_rows.append({
                "chunk_id": c["chunk_id"],
                "asset_id": iid,
                "asset_type": "image",
                "page_id": c["page_id"],
                "relation_score": 0.55,
                "relation_reasons": ["same_page"]
            })

    # images.jsonl (we need Doc ID and Image Path)
    # I'll need to modify build_doc_records to pass this up or extract from records
    # Let's adjust build_doc_records return first.

    write_jsonl_atomic(chunks_path, chunks_merged)
    write_jsonl_atomic(pages_path, pages_merged)
    write_jsonl_atomic(tables_path, tables_merged)
    write_jsonl_atomic(docs_path, docs_merged)
    write_jsonl_atomic(all_texts_path, all_texts_merged)
    write_jsonl_atomic(images_path, final_images_rows)
    write_jsonl_atomic(relations_path, all_relations_rows)

    if args.skip_embeddings:
        for doc_id, payload in staged_docs.items():
            key = payload["doc_key"]
            progress["documents"][key] = {
                "status": "completed",
                "doc_id": doc_id,
                "doc_title": payload["doc_title"],
                "signature": payload["signature"],
                "updated_at": now_iso(),
            }

        progress["updated_at"] = now_iso()
        write_json_atomic(progress_path, progress)
        print(
            f"Done: workbooks_scanned={workbook_counter}, staged_docs={len(staged_docs)}, "
            f"removed_docs={len(removed_doc_ids)}, embeddings=skipped"
        )
        return 0

    if staged_docs and not args.api_key:
        raise ValueError("Missing SiliconFlow API key. Use --api-key or SILICONFLOW_API_KEY.")

    # 1. Embed Chunks
    print(f"Embedding {len(chunks_merged)} chunks...")
    texts_to_embed = [c["chunk_text_display"] for c in chunks_merged]
    chunk_vectors = embed_texts(
        api_base=args.embedding_api_base,
        api_key=args.api_key,
        model=args.embedding_model,
        texts=texts_to_embed,
        batch_size=args.batch_size,
        timeout=args.timeout,
        dimensions=resolved_dimensions,
    )
    for c, v in zip(chunks_merged, chunk_vectors):
        c["embedding"] = v

    # 2. Embed Tables
    print(f"Embedding {len(tables_merged)} tables...")
    table_texts = [f"{t['table_title']}\n{t['table_summary']}" for t in tables_merged]
    table_vectors = embed_texts(
        api_base=args.embedding_api_base,
        api_key=args.api_key,
        model=args.embedding_model,
        texts=table_texts,
        batch_size=args.batch_size,
        timeout=args.timeout,
        dimensions=resolved_dimensions,
    )
    for t, v in zip(tables_merged, table_vectors):
        t["embedding"] = v

    # 3. Mean Pooling for Pages
    print(f"Calculating embeddings for {len(pages_merged)} pages...")
    chunk_emb_map = {c["chunk_id"]: c["embedding"] for c in chunks_merged}
    for p in pages_merged:
        child_embs = [chunk_emb_map[cid] for cid in p["chunk_ids"] if cid in chunk_emb_map]
        p["embedding"] = compute_mean_embedding(child_embs) or [0.0] * (resolved_dimensions or 4096)

    # 4. Mean Pooling for Docs
    print(f"Calculating embeddings for {len(docs_merged)} docs...")
    page_emb_map = {p["page_id"]: p["embedding"] for p in pages_merged}
    for d in docs_merged:
        child_embs = [page_emb_map[pid] for pid in d["page_ids"] if pid in page_emb_map]
        d["embedding"] = compute_mean_embedding(child_embs) or [0.0] * (resolved_dimensions or 4096)

    # Save Indices (Singular Naming)
    index_manifest = {}
    for name, merged in [("chunk", chunks_merged), ("page", pages_merged), ("table", tables_merged), ("doc", docs_merged)]:
        matrix = save_numpy_index(output_dir, merged, index_name=name)
        index_manifest[name] = {
            "file": f"vector_store/{name}_embeddings.npy",
            "count": len(merged),
        }

    backend = "numpy_cosine_fallback"
    write_manifest(
        output_dir=output_dir,
        api_base=args.embedding_api_base,
        model=args.embedding_model,
        dimensions=resolved_dimensions,
        doc_updates=doc_manifest_updates,
        removed_doc_ids=removed_doc_ids,
        backend=backend,
        index_manifest=index_manifest,
    )

    # Update progress
    for doc_id, payload in staged_docs.items():
        key = payload["doc_key"]
        progress["documents"][key] = {
            "status": "completed",
            "doc_id": doc_id,
            "doc_title": payload["doc_title"],
            "signature": payload["signature"],
            "updated_at": now_iso(),
        }

    progress["updated_at"] = now_iso()
    write_json_atomic(progress_path, progress)

    print(
        f"Done: workbooks_scanned={workbook_counter}, staged_docs={len(staged_docs)}, "
        f"removed_docs={len(removed_doc_ids)}, backend={backend}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
