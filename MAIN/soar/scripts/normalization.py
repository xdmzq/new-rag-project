#!/usr/bin/env python3
"""Normalize source_data asset names and support resumable execution.

Behavior:
1. Discover workbook(s) under source_data, excluding temporary Excel files.
2. Move image assets into images/ and rename them to images/figXX.png.
3. Use worksheet 1 column A as the canonical table order and rename markdown files to markdown_tables/tableXX.md.
4. Rename the workbook itself to a canonical annotation filename.
5. Rewrite worksheet column A values to the normalized names.
6. Persist per-workbook progress to JSON so interrupted runs can resume safely.

Default mode is dry-run. Pass --apply to modify files in place.
"""

from __future__ import annotations

import argparse
import base64
import concurrent.futures
import hashlib
import json
import os
import re
import sys
import threading
import time
import uuid
from collections import deque
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

import requests
from openpyxl import load_workbook
from PIL import Image

from env_utils import resolve_env_value, resolve_dashscope_key


DEFAULT_DATA_ROOT = Path("source_data")
DEFAULT_PROGRESS_FILE = DEFAULT_DATA_ROOT / "normalization_progress.json"
LOGIC_VERSION = 6
IMAGE_DIR_NAME = "images"
TABLE_DIR_NAME = "markdown_tables"
WORKBOOK_BASENAME = "标注"
TEMP_PREFIXES = ("~$",)
WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp"}
KNOWN_SUFFIXES = (".png", ".jpg", ".jpeg", ".webp", ".md")

# Table asset filename patterns: files that are table images, not regular images
# These should be moved to markdown_tables/, not images/
TABLE_FILENAME_PREFIXES = frozenset({
    "zhtt", "wctb", "xt",
    "shyt", "shy",
})
TABLE_FILENAME_MARKERS = frozenset({"表格"})  # files containing "表格" in name


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normalize image/table asset names and align workbook column A values.")
    parser.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT, help="Root directory containing source workbooks.")
    parser.add_argument("--workbook", type=Path, help="Normalize a single workbook instead of scanning the whole data root.")
    parser.add_argument("--apply", action="store_true", help="Apply changes in place. Default is dry-run.")
    parser.add_argument(
        "--progress-file",
        type=Path,
        default=DEFAULT_PROGRESS_FILE,
        help="JSON file used to record progress and resume state.",
    )
    parser.add_argument("--no-resume", action="store_true", help="Ignore existing in-progress state and rebuild plans.")

    # Keyword reconstruction arguments
    parser.add_argument("--rebuild-keywords", action="store_true", help="Enable semantic keyword reconstruction using LLM.")
    parser.add_argument("--api-key", default="", help="SiliconFlow API key.")
    parser.add_argument("--keyword-model", default="qwen3.6-plus", help="LLM model for keyword reconstruction.")
    parser.add_argument("--keyword-rpm", type=int, default=500, help="Rate limit (RPM) for keyword API.")
    parser.add_argument("--keyword-max-workers", type=int, default=8, help="Max parallel workers for keyword generation.")
    parser.add_argument("--keyword-timeout", type=float, default=120.0, help="API timeout in seconds.")
    parser.add_argument("--keyword-force", action="store_true", help="Force rebuild even if keywords exist.")
    parser.add_argument(
        "--keyword-sheets",
        default="all",
        help="Comma-separated subset of sheets to rebuild: image,table,text,all. Default: all.",
    )

    args = parser.parse_args()
    args.api_key = resolve_env_value("SILICONFLOW_API_KEY", args.api_key)
    args.dashscope_api_key = resolve_dashscope_key()
    raw_sheet_tokens = [item.strip().lower() for item in str(args.keyword_sheets or "all").split(",") if item.strip()]
    if not raw_sheet_tokens:
        raw_sheet_tokens = ["all"]
    allowed_sheet_tokens = {"all", "image", "table", "text"}
    invalid_tokens = [item for item in raw_sheet_tokens if item not in allowed_sheet_tokens]
    if invalid_tokens:
        raise ValueError(f"Invalid --keyword-sheets values: {invalid_tokens}. Allowed: {sorted(allowed_sheet_tokens)}")
    args.keyword_sheet_set = {"image", "table", "text"} if "all" in raw_sheet_tokens else set(raw_sheet_tokens)
    return args


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_slashes(value: str) -> str:
    return str(value or "").replace("\\", "/")


def normalize_text(text: str) -> str:
    return re.sub(r"\n{3,}", "\n\n", str(text or "").replace("\r\n", "\n").replace("\r", "\n")).strip()


def sanitize_for_excel(text: str) -> str:
    """Remove control characters that openpyxl cannot handle."""
    if not isinstance(text, str):
        return str(text)
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", "", text)


def split_keywords(text: str) -> list[str]:
    parts = re.split(r"[、,，;；|\n]+", str(text or "").strip())
    return [part.strip() for part in parts if part and part.strip()]


def normalize_keyword_output(text: str, *, minimum: int = 3, maximum: int = 6) -> str:
    if str(text).startswith("ERROR"):
        return str(text)
    seen: set[str] = set()
    items: list[str] = []
    for part in split_keywords(text):
        cleaned = re.sub(r"^[\-•\d\.\s]+", "", part).strip()
        cleaned = cleaned.strip("。；;，,：:")
        if len(cleaned) <= 1 or cleaned in seen:
            continue
        seen.add(cleaned)
        items.append(cleaned)
    if len(items) > maximum:
        items = items[:maximum]
    return "、".join(items[:maximum]) if len(items) >= minimum else "、".join(items)


def is_table_asset_file(path: Path) -> bool:
    """Check if a file is a table asset (PNG image used as a table source).

    Table assets are identified by:
    - Filename contains "表格" (table) marker
    - Filename starts with known table prefixes (e.g., zhtt, wctb, xt, shyt)

    These should NOT be mixed with regular images in images/ directory.
    """
    name_lower = path.name.lower()
    # Check for "表格" marker in the filename
    if "表格" in path.name:
        return True
    # Check for known table prefixes
    for prefix in TABLE_FILENAME_PREFIXES:
        if name_lower.startswith(prefix):
            return True
    return False


def strip_known_suffix(value: str) -> str:
    lower = value.lower()
    for suffix in KNOWN_SUFFIXES:
        if lower.endswith(suffix):
            return value[: -len(suffix)]
    return value


def canonical_label(value: str) -> str:
    text = normalize_slashes(str(value or "").strip())
    base = text.split("/")[-1]
    base = strip_known_suffix(base).lower()
    parts = re.findall(r"\d+|[^\d]+", base)
    normalized: list[str] = []
    for part in parts:
        if part.isdigit():
            normalized.append(str(int(part)))
        else:
            normalized.append(re.sub(r"[\s_\-\.]+", "", part))
    return "".join(normalized)


def numeric_signature(value: str) -> tuple[int, ...]:
    text = normalize_slashes(str(value or "").strip())
    base = text.split("/")[-1]
    base = strip_known_suffix(base)
    return tuple(int(part) for part in re.findall(r"\d+", base))


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(path)


def load_progress(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"version": 1, "updated_at": now_iso(), "workbooks": {}}
    return json.loads(path.read_text(encoding="utf-8"))


def save_progress(path: Path, payload: dict[str, Any]) -> None:
    payload["updated_at"] = now_iso()
    write_json(path, payload)


def iter_workbooks(data_root: Path) -> list[Path]:
    workbooks: list[Path] = []
    for path in sorted(data_root.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith(TEMP_PREFIXES):
            continue
        if path.suffix.lower() in WORKBOOK_SUFFIXES:
            workbooks.append(path.resolve())
    return workbooks


def natural_sort_key(path: Path) -> tuple[Any, ...]:
    base = strip_known_suffix(path.name).lower()
    parts = re.findall(r"\d+|[^\d]+", base)
    key: list[Any] = []
    for part in parts:
        if part.isdigit():
            key.append((0, int(part)))
        else:
            key.append((1, re.sub(r"[\s_\-\.]+", "", part)))
    return tuple(key)


def relative_asset_value(source_dir: Path, path: Path) -> str:
    try:
        return normalize_slashes(str(path.resolve().relative_to(source_dir.resolve())))
    except ValueError:
        return normalize_slashes(path.name)


class RateLimiter:
    def __init__(self, rpm: int) -> None:
        if rpm <= 0:
            raise ValueError("rpm must be > 0")
        self.rpm = rpm
        self.window = deque()
        self.lock = threading.Lock()

    def wait(self) -> None:
        while True:
            with self.lock:
                now = time.monotonic()
                cutoff = now - 60.0
                while self.window and self.window[0] < cutoff:
                    self.window.popleft()
                if len(self.window) < self.rpm:
                    self.window.append(now)
                    return
                sleep_for = max(0.05, 60.0 - (now - self.window[0]))
            time.sleep(min(sleep_for, 1.0))


def text_hash(text: str) -> str:
    return hashlib.sha1(text.encode("utf-8")).hexdigest()


class ContextManager:
    """Manages page-level context for keyword reconstruction."""

    def __init__(self) -> None:
        self.context_map: dict[tuple[str, str, str, str], dict[str, Any]] = {}

    def _normalize_key_part(self, value: str) -> str:
        return re.sub(r"[\s_\-\.]+", "", str(value or "").strip()).lower()

    def index_workbook(self, wb) -> None:
        """Index Worksheet 3 (文本) to provide context for other worksheets."""
        if len(wb.worksheets) < 3:
            return

        ws = wb.worksheets[2]
        header = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}
        
        required = ["文件名", "页码", "文本内容"]
        for req in required:
            if req not in header:
                return

        directory_col = header.get("目录")
        file_type_col = header.get("文件类型")
        file_col = header["文件名"]
        page_col = header["页码"]
        text_col = header["文本内容"]
        keyword_col = header.get("关键词")

        for row_idx in range(2, ws.max_row + 1):
            file_raw = ws.cell(row=row_idx, column=file_col).value
            page_raw = ws.cell(row=row_idx, column=page_col).value
            text_raw = ws.cell(row=row_idx, column=text_col).value
            
            if not file_raw or not page_raw or not text_raw:
                continue
                
            dir_key = self._normalize_key_part(ws.cell(row=row_idx, column=directory_col).value or "") if directory_col else ""
            type_key = self._normalize_key_part(ws.cell(row=row_idx, column=file_type_col).value or "") if file_type_col else ""
            file_key = self._normalize_key_part(file_raw)
            page_key = str(page_raw).strip()
            
            self.context_map[(dir_key, type_key, file_key, page_key)] = {
                "text": normalize_text(text_raw),
                "keywords": str(ws.cell(row=row_idx, column=keyword_col).value or "") if keyword_col else ""
            }

    def get_context(self, directory: str, file_type: str, file_name: str, page_no: str) -> dict[str, Any] | None:
        dir_key = self._normalize_key_part(directory)
        type_key = self._normalize_key_part(file_type)
        file_key = self._normalize_key_part(file_name)
        page_key = str(page_no).strip()
        return self.context_map.get((dir_key, type_key, file_key, page_key))


class KeywordRebuilder:
    """Orchestrates keyword reconstruction via SiliconFlow."""

    def __init__(self, dashscope_key: str, siliconflow_key: str, model: str, rpm: int, timeout: float):
        self.dashscope_key = dashscope_key
        self.siliconflow_key = siliconflow_key
        self.model = model
        self.limiter = RateLimiter(rpm)
        self.timeout = timeout
        self.api_url = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
        self.vision_api_url = "https://api.siliconflow.cn/v1/chat/completions"
        self.session = requests.Session()

    def _request(self, messages: list[dict[str, Any]], use_vision: bool = False) -> str:
        payload = {
            "model": "Qwen/Qwen3.5-397B-A17B" if use_vision else self.model,
            "messages": messages,
            "temperature": 0.1,
            "max_tokens": 256,
            "stream": False,
        }
        headers = {
            "Authorization": f"Bearer {self.siliconflow_key if use_vision else self.dashscope_key}",
            "Content-Type": "application/json"
        }
        
        for attempt in range(3):
            self.limiter.wait()
            try:
                response = self.session.post(
                    self.vision_api_url if use_vision else self.api_url,
                    headers=headers, json=payload, timeout=self.timeout
                )
                if response.status_code == 200:
                    data = response.json()
                    content = data["choices"][0]["message"]["content"].strip()
                    normalized = normalize_keyword_output(content)
                    return normalized or "ERROR: empty keyword response"
                if response.status_code in {429, 500, 502, 503, 504}:
                    time.sleep(2**attempt)
                    continue
                return f"ERROR: {response.status_code} {response.text[:300]}"
            except Exception as e:
                if attempt == 2:
                    return f"ERROR: {str(e)}"
                time.sleep(1)
        return "ERROR: Max retries exceeded"

    def rebuild_text(self, text: str) -> str:
        system_prompt = (
            "你是一个专业的投研资料编目助手。请重建适合RAG混合检索的4到6个核心关键词。\n"
            "原则：\n"
            "1. 必须是中文短语，使用顿号“、”分隔。\n"
            "2. 每个关键词不少于6个中文字符，必须包含定语修饰成分，例如“商业航天运载火箭发射”而不是“运载火箭”。\n"
            "3. 优先输出行业词+限定词、产品词+场景词、技术词+应用领域、指标词+度量对象，避免整句标题或泛化短词。\n"
            "4. 不要输出年份区间、“市场情况”、“行业趋势”、“电机介绍”这类泛标题。\n"
            "5. 关键词应兼顾页面主题和检索区分度。\n"
            "6. 只返回关键词列表，不要任何解释。"
        )
        user_content = f"页面正文：\n{normalize_text(text)[:8000]}"
        return self._request([
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content}
        ])

    def rebuild_table(self, table_md: str, context: str) -> str:
        system_prompt = (
            "你是一个专业的投研资料编目助手。根据Markdown表格及其所在页面正文，重建适合RAG混合检索的4到6个核心关键词。\n"
            "原则：\n"
            "1. 每个关键词不少于6个中文字符，必须反映表格主体并包含定语，如“全球卫星互联网市场规模”而不是“市场规模”。\n"
            "2. 优先短标签+限定词，不要照抄整行表题。\n"
            "3. 使用中文顿号“、”分隔。\n"
            "4. 只返回关键词列表，不要任何解释。"
        )
        user_content = f"页面正文：\n{normalize_text(context)[:2400]}\n\n表格内容：\n{normalize_text(table_md)[:6000]}"
        return self._request([
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content}
        ])

    def rebuild_image(self, image_path: Path, context: str) -> str:
        # For Qwen Vision, we need to encode image
        try:
            with open(image_path, "rb") as f:
                img_b64 = base64.b64encode(f.read()).decode("utf-8")
        except Exception as e:
            return f"ERROR_IMAGE_READ: {e}"

        system_prompt = (
            "你是一个专业的投研资料编目助手。请分析图片内容并结合所在页面正文，重建适合RAG混合检索的4到6个核心关键词。\n"
            "原则：\n"
            "1. 每个关键词不少于6个中文字符，要落到图像主体对象+应用场景、技术主题+具体领域或指标主题+度量对象。\n"
            "2. 避免输出泛标题或整句说明，必须包含定语修饰。\n"
            "3. 使用中文顿号“、”分隔。\n"
            "4. 只返回关键词列表，不要任何解释。"
        )
        
        # Multi-modal payload - use vision model on SiliconFlow
        messages = [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": f"上下文背景：\n{context[:2000]}"},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{img_b64}", "detail": "low"}
                    }
                ]
            }
        ]
        return self._request(messages, use_vision=True)


def enumerate_files(source_dir: Path, *, image_files: bool) -> list[Path]:
    paths: list[Path] = []
    if image_files:
        candidates: list[Path] = []
        image_dir = source_dir / IMAGE_DIR_NAME
        candidates.extend(source_dir.glob("*"))
        if image_dir.exists():
            candidates.extend(image_dir.glob("*"))
        for path in sorted(candidates, key=natural_sort_key):
            if not path.is_file():
                continue
            if path.suffix.lower() not in IMAGE_SUFFIXES:
                continue
            # Exclude table asset files from image enumeration
            if is_table_asset_file(path):
                continue
            resolved = path.resolve()
            if resolved not in paths:
                paths.append(resolved)
        return paths

    # For tables: first check markdown_tables/ for .md files (post-conversion state)
    table_dir = source_dir / TABLE_DIR_NAME
    if table_dir.exists():
        for path in sorted(table_dir.glob("*.md"), key=natural_sort_key):
            if path.is_file():
                paths.append(path.resolve())
        return paths

    # Fallback: enumerate table PNG assets from root directory (pre-conversion state)
    # These are PNG files that are actually tables (identified by filename pattern)
    for path in sorted(source_dir.glob("*"), key=natural_sort_key):
        if not path.is_file():
            continue
        if path.suffix.lower() not in IMAGE_SUFFIXES:
            continue
        # Only include known table asset files
        if is_table_asset_file(path):
            resolved = path.resolve()
            if resolved not in paths:
                paths.append(resolved)
    return paths


def resolve_asset_path(source_dir: Path, raw_value: str, *, image_files: bool) -> Path | None:
    raw = str(raw_value or "").strip()
    if not raw:
        return None

    search_space = enumerate_files(source_dir, image_files=image_files)
    raw_path = Path(normalize_slashes(raw))

    direct_candidates: list[Path] = []
    if image_files:
        image_dir = source_dir / IMAGE_DIR_NAME
        if raw_path.suffix.lower() in IMAGE_SUFFIXES:
            direct_candidates.extend([source_dir / raw_path, source_dir / raw_path.name, image_dir / raw_path.name])
        else:
            direct_candidates.extend([source_dir / raw_path.name, image_dir / raw_path.name])
            for suffix in sorted(IMAGE_SUFFIXES):
                direct_candidates.append(source_dir / f"{raw_path.name}{suffix}")
                direct_candidates.append(image_dir / f"{raw_path.name}{suffix}")
    else:
        table_dir = source_dir / TABLE_DIR_NAME
        if raw_path.suffix.lower() == ".md":
            direct_candidates.extend([source_dir / raw_path, table_dir / raw_path.name])
        else:
            direct_candidates.extend([table_dir / f"{raw_path.name}.md", source_dir / f"{raw_path.name}.md"])

    for candidate in direct_candidates:
        if candidate.exists():
            return candidate.resolve()

    wanted = canonical_label(raw)
    exact_matches = [
        candidate
        for candidate in search_space
        if canonical_label(candidate.name) == wanted or canonical_label(candidate.stem) == wanted
    ]
    if len(exact_matches) == 1:
        return exact_matches[0]
    if exact_matches:
        return exact_matches[0]

    wanted_sig = numeric_signature(raw)
    if wanted_sig:
        sig_matches = [candidate for candidate in search_space if numeric_signature(candidate.name) == wanted_sig]
        if len(sig_matches) == 1:
            return sig_matches[0]
    return None


def resolve_output_path(source_dir: Path, sheet_value: str) -> Path:
    """Resolve the output path for a table asset.

    Tables go into markdown_tables/ as .md files.
    Uses the stem of the sheet value to construct the filename.
    """
    raw = str(sheet_value or "").strip()
    raw_path = Path(raw)
    # If already a markdown path, resolve it directly
    if raw_path.suffix.lower() == ".md":
        return (source_dir / raw_path).resolve()
    # Use the stem to construct markdown_tables/<stem>.md
    stem = raw_path.stem if raw_path.suffix else raw_path.name
    table_dir = source_dir / TABLE_DIR_NAME
    return (table_dir / f"{stem}.md").resolve()


def collect_rows(ws) -> list[int]:
    rows: list[int] = []
    for row_index in range(2, ws.max_row + 1):
        value = ws.cell(row=row_index, column=1).value
        if value is None or str(value).strip() == "":
            continue
        rows.append(row_index)
    return rows


def target_value_for(kind: str, seq: int, width: int) -> str:
    if kind == "image":
        return f"{IMAGE_DIR_NAME}/fig{seq:0{width}d}.png"
    return f"{TABLE_DIR_NAME}/table{seq:0{width}d}.md"


def canonical_workbook_name(workbook_path: Path) -> str:
    suffix = workbook_path.suffix.lower() or ".xlsx"
    return f"{WORKBOOK_BASENAME}{suffix}"


def make_plan_entry(
    *,
    kind: str,
    row_index: int,
    current_value: str,
    source_path: Path,
    target_value: str,
    target_path: Path,
    convert_to_png: bool,
) -> dict[str, Any]:
    return {
        "kind": kind,
        "row_index": row_index,
        "current_value": current_value,
        "source_path": str(source_path.resolve()),
        "target_value": target_value,
        "target_path": str(target_path.resolve()),
        "convert_to_png": convert_to_png,
        "stage": "pending",
        "temp_path": "",
    }


def build_asset_plans(source_dir: Path, *, kind: str) -> list[dict[str, Any]]:
    candidates = enumerate_files(source_dir, image_files=(kind == "image"))
    width = max(2, len(str(max(len(candidates), 1))))
    plans: list[dict[str, Any]] = []

    for seq, source_path in enumerate(candidates, start=1):
        target_value = target_value_for(kind, seq, width)
        target_path = source_dir / target_value
        current_value = relative_asset_value(source_dir, source_path)
        plans.append(
            make_plan_entry(
                kind=kind,
                row_index=0,
                current_value=current_value,
                source_path=source_path,
                target_value=target_value,
                target_path=target_path,
                convert_to_png=(kind == "image" and source_path.suffix.lower() != ".png"),
            )
        )
    return plans


def plan_aliases(plan: dict[str, Any]) -> set[str]:
    source_path = Path(plan["source_path"])
    target_path = Path(plan["target_path"])
    aliases = {
        normalize_slashes(str(plan["current_value"])),
        normalize_slashes(source_path.name),
        normalize_slashes(source_path.stem),
        normalize_slashes(str(plan["target_value"])),
        normalize_slashes(target_path.name),
        normalize_slashes(target_path.stem),
    }
    if plan["kind"] == "image":
        aliases.add(normalize_slashes(f"{IMAGE_DIR_NAME}/{source_path.name}"))
        aliases.add(normalize_slashes(f"{IMAGE_DIR_NAME}/{target_path.name}"))
    if plan["kind"] == "table":
        aliases.add(normalize_slashes(f"{TABLE_DIR_NAME}/{source_path.name}"))
        aliases.add(normalize_slashes(f"{TABLE_DIR_NAME}/{target_path.name}"))
    return {alias for alias in aliases if alias}


def resolve_plan_for_value(raw_value: str, plans: list[dict[str, Any]]) -> dict[str, Any] | None:
    raw = normalize_slashes(str(raw_value or "").strip())
    if not raw:
        return None

    exact_matches = [plan for plan in plans if raw in plan_aliases(plan)]
    if len(exact_matches) == 1:
        return exact_matches[0]
    if exact_matches:
        return exact_matches[0]

    wanted = canonical_label(raw)
    canonical_matches = [plan for plan in plans if any(canonical_label(alias) == wanted for alias in plan_aliases(plan))]
    if len(canonical_matches) == 1:
        return canonical_matches[0]
    if canonical_matches:
        return canonical_matches[0]

    wanted_sig = numeric_signature(raw)
    if wanted_sig:
        signature_matches = [plan for plan in plans if any(numeric_signature(alias) == wanted_sig for alias in plan_aliases(plan))]
        if len(signature_matches) == 1:
            return signature_matches[0]
    return None


def build_sheet_updates(source_dir: Path, ws, *, kind: str, plans: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    updates: list[dict[str, Any]] = []
    deleted_missing = 0

    for row_index in collect_rows(ws):
        current_value = str(ws.cell(row=row_index, column=1).value).strip()
        plan = resolve_plan_for_value(current_value, plans)
        if plan is None:
            updates.append(
                {
                    "row_index": row_index,
                    "current_value": current_value,
                    "target_value": "",
                    "action": "delete_missing",
                }
            )
            deleted_missing += 1
            continue
        updates.append(
            {
                "row_index": row_index,
                "current_value": current_value,
                "target_value": plan["target_value"],
                "action": "align",
            }
        )
    return updates, deleted_missing


def build_workbook_entry(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path)
    source_dir = workbook_path.parent
    workbook_target_path = source_dir / canonical_workbook_name(workbook_path)

    image_plans = build_asset_plans(source_dir, kind="image")
    table_plans = build_asset_plans(source_dir, kind="table")
    image_updates: list[dict[str, Any]] = []
    table_updates: list[dict[str, Any]] = []
    warnings: list[str] = []
    deleted_missing_image_rows = 0
    deleted_missing_table_rows = 0

    if wb.worksheets:
        updates, deleted_missing = build_sheet_updates(source_dir, wb.worksheets[0], kind="image", plans=image_plans)
        image_updates = updates
        deleted_missing_image_rows = deleted_missing

    if len(wb.worksheets) >= 2:
        updates, deleted_missing = build_sheet_updates(source_dir, wb.worksheets[1], kind="table", plans=table_plans)
        table_updates = updates
        deleted_missing_table_rows = deleted_missing

    all_plans = image_plans + table_plans
    changed_images = sum(1 for plan in image_plans if plan["current_value"] != plan["target_value"])
    changed_tables = sum(1 for plan in table_plans if plan["current_value"] != plan["target_value"])

    return {
        "logic_version": LOGIC_VERSION,
        "status": "planned",
        "workbook_path": str(workbook_path.resolve()),
        "workbook_target_path": str(workbook_target_path.resolve()),
        "source_dir": str(source_dir.resolve()),
        "planned_at": now_iso(),
        "started_at": "",
        "finished_at": "",
        "warnings": warnings,
        "workbook_name_to_align": int(workbook_path.resolve() != workbook_target_path.resolve()),
        "image_files_to_align": changed_images,
        "table_files_to_align": changed_tables,
        "image_rows_to_align": sum(1 for update in image_updates if update["current_value"] != update["target_value"]),
        "table_rows_to_align": sum(1 for update in table_updates if update["current_value"] != update["target_value"]),
        "deleted_missing_image_rows": deleted_missing_image_rows,
        "deleted_missing_table_rows": deleted_missing_table_rows,
        "plans": all_plans,
        "image_sheet_updates": image_updates,
        "table_sheet_updates": table_updates,
        "error": "",
    }


def reconcile_plan(plan: dict[str, Any]) -> None:
    source_path = Path(plan["source_path"])
    target_path = Path(plan["target_path"])
    temp_value = str(plan.get("temp_path", "") or "")
    temp_path = Path(temp_value) if temp_value else None

    if target_path.exists():
        if temp_path is None or not temp_path.exists():
            if source_path == target_path or not source_path.exists():
                plan["stage"] = "finalized"
                return
    if temp_path is not None and temp_path.exists():
        plan["stage"] = "staged"
        return
    if source_path.exists():
        plan["stage"] = "pending"


def stage_pending_moves(entry: dict[str, Any], progress: dict[str, Any], progress_path: Path) -> None:
    for plan in entry["plans"]:
        reconcile_plan(plan)
        if plan["stage"] != "pending":
            continue
        if plan["convert_to_png"]:
            continue

        source_path = Path(plan["source_path"])
        target_path = Path(plan["target_path"])
        if source_path == target_path:
            plan["stage"] = "finalized"
            save_progress(progress_path, progress)
            continue

        temp_path = source_path.with_name(f".__normalize_tmp__{uuid.uuid4().hex}{source_path.suffix.lower()}")
        source_path.replace(temp_path)
        plan["temp_path"] = str(temp_path)
        plan["stage"] = "staged"
        save_progress(progress_path, progress)


def finalize_plan(entry: dict[str, Any], progress: dict[str, Any], progress_path: Path) -> None:
    for plan in entry["plans"]:
        reconcile_plan(plan)
        if plan["stage"] == "finalized":
            continue

        source_path = Path(plan["source_path"])
        target_path = Path(plan["target_path"])
        target_path.parent.mkdir(parents=True, exist_ok=True)

        if plan["convert_to_png"]:
            if target_path.exists() and not source_path.exists():
                plan["stage"] = "finalized"
                save_progress(progress_path, progress)
                continue
            with Image.open(source_path) as image:
                rendered = image.convert("RGB") if image.mode not in {"RGB", "L"} else image
                rendered.save(target_path, format="PNG")
            if source_path != target_path and source_path.exists():
                source_path.unlink()
            plan["stage"] = "finalized"
            save_progress(progress_path, progress)
            continue

        temp_value = str(plan.get("temp_path", "") or "")
        working_path = Path(temp_value) if temp_value else source_path
        if working_path == target_path:
            plan["stage"] = "finalized"
            save_progress(progress_path, progress)
            continue
        working_path.replace(target_path)
        plan["stage"] = "finalized"
        save_progress(progress_path, progress)


def apply_sheet_updates(ws, updates: list[dict[str, Any]], header_value: str) -> None:
    ws.cell(row=1, column=1).value = header_value
    for update in updates:
        value = update["target_value"]
        ws.cell(row=int(update["row_index"]), column=1).value = None if value in {"", None} else sanitize_for_excel(str(value))


def save_workbook_alignment(workbook_path: Path, entry: dict[str, Any]) -> None:
    wb = load_workbook(workbook_path)
    if wb.worksheets:
        apply_sheet_updates(wb.worksheets[0], entry.get("image_sheet_updates", []), "图片")
    if len(wb.worksheets) >= 2:
        apply_sheet_updates(wb.worksheets[1], entry.get("table_sheet_updates", []), "Markdown文件")
    wb.save(workbook_path)


def apply_keyword_updates(ws, updates: list[dict[str, Any]], column_name: str = "关键词") -> None:
    header = {str(cell.value).strip().strip("'"): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}
    if column_name not in header:
        print(f"  Warning: Column {column_name} not found in {ws.title}. Header keys: {list(header.keys())}")
        return
    col_idx = header[column_name]
    print(f"  Writing {len(updates)} keywords to {ws.title} at column {col_idx}...")
    for update in updates:
        row_idx = int(update["row_index"])
        new_val = update.get("new_keywords")
        if new_val and not str(new_val).startswith("ERROR"):
            ws.cell(row=row_idx, column=col_idx).value = sanitize_for_excel(new_val)
        if row_idx <= 5:
            print(f"    - Row {row_idx}: {str(new_val)[:50]}...")


def rebuild_keywords_for_workbook(
    workbook_path: Path,
    entry: dict[str, Any],
    args: argparse.Namespace,
    progress: dict[str, Any],
    progress_path: Path,
) -> None:
    wb = load_workbook(workbook_path)
    context_mgr = ContextManager()
    context_mgr.index_workbook(wb)
    
    rebuilder = KeywordRebuilder(
        dashscope_key=args.dashscope_api_key,
        siliconflow_key=args.api_key,
        model=args.keyword_model,
        rpm=args.keyword_rpm,
        timeout=args.keyword_timeout
    )
    
    selected_sheets = getattr(args, "keyword_sheet_set", {"image", "table", "text"})

    # 1. Prepare Text jobs
    text_jobs = []
    if "text" in selected_sheets and len(wb.worksheets) >= 3:
        ws = wb.worksheets[2]
        header = {str(cell.value).strip().strip("'"): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}
        if "文本内容" in header and "关键词" in header:
            text_col = header["文本内容"]
            keyword_col = header["关键词"]
            for row_idx in range(2, ws.max_row + 1):
                text = str(ws.cell(row=row_idx, column=text_col).value or "").strip()
                if not text: continue
                existing = str(ws.cell(row=row_idx, column=keyword_col).value or "").strip()
                if not args.keyword_force and existing: continue
                text_jobs.append({"row_index": row_idx, "text": text, "kind": "text"})

    # 2. Prepare Table jobs
    table_jobs = []
    if "table" in selected_sheets and len(wb.worksheets) >= 2:
        ws = wb.worksheets[1]
        header = {str(cell.value).strip().strip("'"): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}
        if "Markdown文件" in header and "关键词" in header:
            file_col = header["Markdown文件"]
            keyword_col = header["关键词"]
            file_name_col = header.get("文件名", 1) # Fallback to col 1 if not found
            page_col = header.get("页码", 1) 

            for row_idx in range(2, ws.max_row + 1):
                rel_path = str(ws.cell(row=row_idx, column=file_col).value or "").strip()
                if not rel_path: continue
                existing = str(ws.cell(row=row_idx, column=keyword_col).value or "").strip()
                if not args.keyword_force and existing: continue
                
                md_path = workbook_path.parent / rel_path
                table_md = md_path.read_text(encoding="utf-8") if md_path.exists() else ""
                
                # Context lookup
                fn = str(ws.cell(row=row_idx, column=file_name_col).value or "") if "文件名" in header else Path(rel_path).stem
                pg = str(ws.cell(row=row_idx, column=page_col).value or "")
                directory = str(ws.cell(row=row_idx, column=header["目录"]).value or "") if "目录" in header else ""
                file_type = str(ws.cell(row=row_idx, column=header["文件类型"]).value or "") if "文件类型" in header else ""
                ctx = context_mgr.get_context(directory, file_type, fn, pg)
                context_text = ctx["text"] if ctx else ""
                
                table_jobs.append({"row_index": row_idx, "table_md": table_md, "context": context_text, "kind": "table"})

    # 3. Prepare Image jobs
    image_jobs = []
    if "image" in selected_sheets and len(wb.worksheets) >= 1:
        ws = wb.worksheets[0]
        header = {str(cell.value).strip().strip("'"): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}
        if "图片" in header and "关键词" in header:
            file_col = header["图片"]
            keyword_col = header["关键词"]
            file_name_col = header.get("文件名", 1)
            page_col = header.get("页码", 1)

            for row_idx in range(2, ws.max_row + 1):
                rel_path = str(ws.cell(row=row_idx, column=file_col).value or "").strip()
                if not rel_path: continue
                existing = str(ws.cell(row=row_idx, column=keyword_col).value or "").strip()
                if not args.keyword_force and existing: continue
                
                img_path = workbook_path.parent / rel_path
                if not img_path.exists(): continue
                
                fn = str(ws.cell(row=row_idx, column=file_name_col).value or "") if "文件名" in header else Path(rel_path).stem
                pg = str(ws.cell(row=row_idx, column=page_col).value or "")
                directory = str(ws.cell(row=row_idx, column=header["目录"]).value or "") if "目录" in header else ""
                file_type = str(ws.cell(row=row_idx, column=header["文件类型"]).value or "") if "文件类型" in header else ""
                ctx = context_mgr.get_context(directory, file_type, fn, pg)
                context_text = ctx["text"] if ctx else ""
                
                image_jobs.append({"row_index": row_idx, "image_path": img_path, "context": context_text, "kind": "image"})

    def process_job(job):
        if job["kind"] == "text":
            res = rebuilder.rebuild_text(job["text"])
        elif job["kind"] == "table":
            res = rebuilder.rebuild_table(job["table_md"], job["context"])
        elif job["kind"] == "image":
            res = rebuilder.rebuild_image(job["image_path"], job["context"])
        else:
            res = "UNKNOWN_KIND"
        return job, res

    all_jobs = text_jobs + table_jobs + image_jobs
    print(f"  Debug: text_jobs={len(text_jobs)} table_jobs={len(table_jobs)} image_jobs={len(image_jobs)}")
    if not all_jobs:
        return

    print(f"  Rebuilding {len(all_jobs)} keywords using {args.keyword_model}...")
    
    updates_text = []
    updates_table = []
    updates_image = []
    error_samples: list[str] = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=args.keyword_max_workers) as executor:
        future_map = {executor.submit(process_job, job): job for job in all_jobs}
        done_count = 0
        for future in concurrent.futures.as_completed(future_map):
            job, result = future.result()
            normalized = normalize_keyword_output(result)
            job["new_keywords"] = normalized
            if str(normalized).startswith("ERROR") and len(error_samples) < 8:
                error_samples.append(f"{job['kind']} row={job['row_index']} -> {normalized}")
            if job["kind"] == "text": updates_text.append(job)
            elif job["kind"] == "table": updates_table.append(job)
            elif job["kind"] == "image": updates_image.append(job)
            
            done_count += 1
            if done_count % 10 == 0 or done_count == len(all_jobs):
                print(f"    Keywords: {done_count}/{len(all_jobs)} completed")

    if error_samples:
        print("  Keyword rebuild warnings:")
        for item in error_samples:
            print(f"    - {item}")

    # Apply updates to memory wb
    sheet_map = {ws.title: ws for ws in wb.worksheets}
    if updates_image and "图片" in sheet_map: 
        apply_keyword_updates(sheet_map["图片"], updates_image)
    if updates_table and "表格" in sheet_map:
        apply_keyword_updates(sheet_map["表格"], updates_table)
    if updates_text and "文本" in sheet_map:
        apply_keyword_updates(sheet_map["文本"], updates_text)
    
    wb.save(workbook_path)


def run_apply_for_workbook(entry: dict[str, Any], progress: dict[str, Any], progress_path: Path, args: argparse.Namespace) -> None:
    original_key = str(entry["workbook_path"])
    entry["status"] = "in_progress"
    if not entry["started_at"]:
        entry["started_at"] = now_iso()
    save_progress(progress_path, progress)

    stage_pending_moves(entry, progress, progress_path)
    finalize_plan(entry, progress, progress_path)
    save_workbook_alignment(Path(entry["workbook_path"]), entry)
    
    if args.rebuild_keywords:
        rebuild_keywords_for_workbook(Path(entry["workbook_path"]), entry, args, progress, progress_path)
        
    rename_workbook_file(entry, progress, progress_path, original_key)

    entry["status"] = "done"
    entry["finished_at"] = now_iso()
    entry["error"] = ""
    save_progress(progress_path, progress)


def rename_workbook_file(
    entry: dict[str, Any],
    progress: dict[str, Any],
    progress_path: Path,
    original_key: str,
) -> str:
    current_path = Path(entry["workbook_path"])
    target_path = Path(entry.get("workbook_target_path") or current_path)
    if current_path == target_path:
        return original_key
    if target_path.exists() and target_path != current_path:
        raise FileExistsError(f"Target workbook already exists: {target_path}")
    current_path.replace(target_path)
    entry["workbook_path"] = str(target_path.resolve())
    progress["workbooks"][entry["workbook_path"]] = entry
    if original_key in progress["workbooks"] and original_key != entry["workbook_path"]:
        del progress["workbooks"][original_key]
    save_progress(progress_path, progress)
    return entry["workbook_path"]


def main() -> int:
    args = parse_args()
    progress_path = args.progress_file.resolve()
    progress = load_progress(progress_path)

    workbooks = [args.workbook.resolve()] if args.workbook else iter_workbooks(args.data_root.resolve())
    if not workbooks:
        print("No workbook found.", file=sys.stderr)
        return 1

    mode = "APPLY" if args.apply else "DRY-RUN"
    print(f"[{mode}] workbooks={len(workbooks)} progress={progress_path}")

    total_image_changes = 0
    total_table_changes = 0
    total_workbook_changes = 0
    total_deleted_missing_rows = 0
    total_errors = 0
    collected_warnings: list[str] = []

    for workbook_path in workbooks:
        key = str(workbook_path)
        entry = progress["workbooks"].get(key)
        entry_logic_version = int(entry.get("logic_version", 0) or 0) if entry else 0

        if args.apply and not args.no_resume and entry_logic_version == LOGIC_VERSION and entry and entry.get("status") in {"in_progress", "error"} and entry.get("plans"):
            workbook_entry = entry
        elif entry and entry_logic_version == LOGIC_VERSION and entry.get("status") == "done" and args.apply and not args.no_resume and not args.rebuild_keywords:
            print(f"- {workbook_path}: already_done")
            total_workbook_changes += int(entry.get("workbook_name_to_align", 0) or 0)
            total_image_changes += int(entry.get("image_files_to_align", entry.get("image_rows_to_align", 0)) or 0)
            total_table_changes += int(entry.get("table_files_to_align", entry.get("table_rows_to_align", 0)) or 0)
            total_deleted_missing_rows += int(entry.get("deleted_missing_image_rows", 0) or 0) + int(entry.get("deleted_missing_table_rows", 0) or 0)
            collected_warnings.extend(entry.get("warnings", []) or [])
            continue
        else:
            workbook_entry = build_workbook_entry(workbook_path)
            progress["workbooks"][key] = workbook_entry
            save_progress(progress_path, progress)

        total_workbook_changes += int(workbook_entry.get("workbook_name_to_align", 0) or 0)
        total_image_changes += int(workbook_entry.get("image_files_to_align", workbook_entry.get("image_rows_to_align", 0)) or 0)
        total_table_changes += int(workbook_entry.get("table_files_to_align", workbook_entry.get("table_rows_to_align", 0)) or 0)
        total_deleted_missing_rows += int(workbook_entry.get("deleted_missing_image_rows", 0) or 0) + int(workbook_entry.get("deleted_missing_table_rows", 0) or 0)
        collected_warnings.extend(workbook_entry.get("warnings", []) or [])

        try:
            if args.apply:
                if args.rebuild_keywords:
                    if not args.api_key:
                        raise ValueError("Missing SiliconFlow API key for keyword reconstruction.")

                run_apply_for_workbook(workbook_entry, progress, progress_path, args)
            print(
                f"- {workbook_path}: "
                f"workbook_name_to_align={workbook_entry.get('workbook_name_to_align', 0)} "
                f"image_files_to_align={workbook_entry.get('image_files_to_align', 0)} "
                f"table_files_to_align={workbook_entry.get('table_files_to_align', 0)} "
                f"image_rows_to_align={workbook_entry['image_rows_to_align']} "
                f"table_rows_to_align={workbook_entry['table_rows_to_align']} "
                f"deleted_missing_rows={int(workbook_entry.get('deleted_missing_image_rows', 0) or 0) + int(workbook_entry.get('deleted_missing_table_rows', 0) or 0)} "
                f"warnings={len(workbook_entry['warnings'])} "
                f"status={workbook_entry['status']}"
            )
        except Exception as exc:  # pragma: no cover
            workbook_entry["status"] = "error"
            workbook_entry["error"] = f"{type(exc).__name__}: {exc}"
            save_progress(progress_path, progress)
            total_errors += 1
            print(f"- {workbook_path}: status=error error={workbook_entry['error']}", file=sys.stderr)

    print(
        f"summary: workbook_name_to_align={total_workbook_changes} "
        f"image_files_to_align={total_image_changes} "
        f"table_files_to_align={total_table_changes} "
        f"deleted_missing_rows={total_deleted_missing_rows} "
        f"errors={total_errors}"
    )
    if collected_warnings:
        print("warnings:")
        for warning in collected_warnings:
            print(f"  - {warning}")
    return 1 if total_errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
