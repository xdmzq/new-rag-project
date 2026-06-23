#!/usr/bin/env python3
"""Incrementally extract document texts into workbook worksheet #3 (`文本`).

This script supports checkpoint resume and incremental updates for newly added
documents. Existing extracted documents are recorded into a progress file so
future runs only process changed/new files.
"""

from __future__ import annotations

import argparse
import base64
import concurrent.futures
import json
import os
import re
import subprocess
import sys
import tempfile
import threading
import time
import zipfile
import docx
from collections import deque
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator, Sequence
from xml.etree import ElementTree as ET

import fitz
import requests
from openpyxl import Workbook, load_workbook

from env_utils import resolve_env_value, resolve_claude_key


DEFAULT_DATA_ROOT = Path("source_data")
DEFAULT_PROGRESS_FILE_NAME = "extract_document_texts_progress.json"
DEFAULT_SHEET_NAME = "文本"
DEFAULT_WORKBOOK_NAME = "标注.xlsx"
IMAGE_SHEET_NAME = "图片"
TABLE_SHEET_NAME = "表格"
DEFAULT_MODEL = "qwen3.6-plus"
DEFAULT_API_BASE = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
DEFAULT_SILICONFLOW_API_BASE = "https://api.siliconflow.cn/v1/chat/completions"
SUPPORTED_SOURCE_SUFFIXES = {".pptx", ".ppt", ".pdf", ".docx", ".doc"}
SUPPORTED_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
TEMP_PREFIXES = ("~$",)

# OCR Defaults
DEFAULT_OCR_MODEL = "Qwen/Qwen3.5-397B-A17B"
DEFAULT_OCR_STRATEGY = "auto"  # auto, fitz, tesseract, siliconflow
DEFAULT_OCR_THRESHOLD = 0.2    # digit-to-total ratio above this triggers OCR

NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
}


@dataclass
class TextRow:
    text_id: str
    keywords: str
    directory: str
    file_type: str
    file_name: str
    page_no: str
    text_content: str


@dataclass(frozen=True)
class WorkbookJob:
    workbook_path: Path
    source_dir: Path


class RateLimiter:
    def __init__(self, rpm: int) -> None:
        if rpm <= 0:
            raise ValueError("rpm must be > 0")
        self.rpm = rpm
        self.window = deque()
        self.lock = threading.Lock()

    def wait(self) -> None:
        while True:
            with self.lock:
                now = time.monotonic()
                cutoff = now - 60.0
                while self.window and self.window[0] < cutoff:
                    self.window.popleft()
                if len(self.window) < self.rpm:
                    self.window.append(now)
                    return
                sleep_for = max(0.05, 60.0 - (now - self.window[0]))
            time.sleep(min(sleep_for, 1.0))


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_whitespace(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def sanitize_for_excel(text: str) -> str:
    """Remove control characters that openpyxl cannot handle."""
    if not isinstance(text, str):
        return str(text)
    # ASCII control characters 0-31, except 9 (tab), 10 (LF), 13 (CR)
    # Also 127 (DEL) and 128-159 (C1 controls)
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", "", text)


def numeric_sort_key(path: str) -> tuple[int, str]:
    match = re.search(r"(\d+)", Path(path).stem)
    number = int(match.group(1)) if match else sys.maxsize
    return number, path


def load_progress(progress_file: Path) -> dict:
    if progress_file.exists():
        try:
            return json.loads(progress_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            pass
    return {"version": 1, "updated_at": now_iso(), "workbooks": {}}


def save_progress(progress_file: Path, progress: dict) -> None:
    progress["updated_at"] = now_iso()
    progress_file.parent.mkdir(parents=True, exist_ok=True)
    temp_path = progress_file.with_suffix(progress_file.suffix + ".tmp")
    temp_path.write_text(json.dumps(progress, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(progress_file)


def resolve_progress_file(data_root: Path, progress_file: Path | None) -> Path:
    if progress_file is not None:
        return progress_file.resolve()
    return (data_root / DEFAULT_PROGRESS_FILE_NAME).resolve()


def file_signature(path: Path) -> dict[str, int]:
    stat = path.stat()
    return {"size": int(stat.st_size), "mtime_ns": int(stat.st_mtime_ns)}


def iter_source_files(source_dir: Path) -> Iterator[Path]:
    for path in sorted(source_dir.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith(TEMP_PREFIXES):
            continue
        if path.suffix.lower() in SUPPORTED_SOURCE_SUFFIXES:
            yield path


def iter_workbooks(data_root: Path) -> Iterator[Path]:
    for path in sorted(data_root.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith(TEMP_PREFIXES):
            continue
        if path.suffix.lower() in SUPPORTED_WORKBOOK_SUFFIXES:
            yield path


def discover_source_dirs(data_root: Path) -> list[Path]:
    source_dirs: set[Path] = set()
    for path in iter_source_files(data_root):
        source_dirs.add(path.parent.resolve())
    return sorted(source_dirs)


def create_annotation_workbook(workbook_path: Path) -> None:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    image_ws = wb.active
    image_ws.title = IMAGE_SHEET_NAME
    table_ws = wb.create_sheet(title=TABLE_SHEET_NAME)
    text_ws = wb.create_sheet(title=DEFAULT_SHEET_NAME)

    image_ws.append(("图片", "关键词", "目录", "文件类型", "文件名", "页码"))
    table_ws.append(("表格", "关键词", "目录", "文件类型", "文件名", "页码"))
    text_ws.append(("文本ID", "关键词", "目录", "文件类型", "文件名", "页码", "文本内容"))
    wb.save(workbook_path)


def discover_or_bootstrap_workbook_jobs(data_root: Path, *, bootstrap_workbooks: bool) -> list[WorkbookJob]:
    if not bootstrap_workbooks:
        return discover_workbook_jobs(data_root)

    existing_by_dir: dict[Path, Path] = {}
    for workbook_path in iter_workbooks(data_root):
        existing_by_dir.setdefault(workbook_path.parent.resolve(), workbook_path.resolve())

    jobs: list[WorkbookJob] = []
    for source_dir in discover_source_dirs(data_root):
        workbook_path = existing_by_dir.get(source_dir)
        if workbook_path is None:
            workbook_path = (source_dir / DEFAULT_WORKBOOK_NAME).resolve()
            create_annotation_workbook(workbook_path)
        jobs.append(WorkbookJob(workbook_path=workbook_path, source_dir=source_dir))
    return jobs


def discover_workbook_jobs(data_root: Path) -> list[WorkbookJob]:
    return [WorkbookJob(workbook_path=p, source_dir=p.parent) for p in iter_workbooks(data_root)]


def doc_key_from_text_id(text_id: str) -> str:
    text = str(text_id or "").strip()
    if "#" in text:
        return text.rsplit("#", 1)[0]
    return text


def row_sort_key(row: TextRow) -> tuple[str, str, str, int, str]:
    page_no = str(row.page_no).strip()
    page_num = int(page_no) if page_no.isdigit() else sys.maxsize
    return row.directory, row.file_name, row.file_type, page_num, row.text_id


def extract_text_from_pptx(path: Path) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    with zipfile.ZipFile(path) as archive:
        slide_names = [
            name
            for name in archive.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        ]
        for index, slide_name in enumerate(sorted(slide_names, key=numeric_sort_key), start=1):
            root = ET.fromstring(archive.read(slide_name))
            parts = [node.text or "" for node in root.findall(".//a:t", NS)]
            rows.append((str(index), normalize_whitespace("\n".join(part for part in parts if part))))
    return rows


def extract_text_from_ppt(path: Path) -> list[tuple[str, str]]:
    catppt = shutil_which("catppt")
    if not catppt:
        raise RuntimeError("catppt is not installed; cannot extract .ppt files")
    result = subprocess.run([catppt, str(path)], check=True, capture_output=True, text=True)
    return [("1", normalize_whitespace(result.stdout))]


def extract_text_from_docx(path: Path) -> list[tuple[str, str]]:
    """Extract sections from DOCX based on headings."""
    try:
        doc = docx.Document(path)
    except Exception as exc:
        print(f"  [WARN] Failed to parse DOCX with python-docx: {exc}", file=sys.stderr)
        return [("1", "")]

    sections: list[tuple[str, list[str]]] = []
    current_title = ""
    current_lines: list[str] = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        
        # Detect headings (heuristic: starts with Heading style or is bold/short)
        is_heading = para.style.name.startswith("Heading") or (
            len(text) < 100 and any(run.bold for run in para.runs)
        )
        
        if is_heading:
            if current_lines:
                sections.append((current_title, current_lines))
                current_lines = []
            current_title = text
        else:
            current_lines.append(text)

    if current_lines or current_title:
        sections.append((current_title, current_lines))

    if not sections:
        return [("1", "")]

    rows: list[tuple[str, str]] = []
    for i, (title, lines) in enumerate(sections, start=1):
        content = "\n".join(lines)
        full_text = f"{title}\n\n{content}" if title else content
        rows.append((str(i), normalize_whitespace(full_text)))
    return rows


def extract_text_from_doc(path: Path) -> list[tuple[str, str]]:
    antiword = shutil_which("antiword")
    if not antiword:
        raise RuntimeError("antiword is not installed; cannot extract .doc files")
    result = subprocess.run([antiword, str(path)], check=True, capture_output=True, text=True)
    return [("1", normalize_whitespace(result.stdout))]


def ocr_pdf_page_siliconflow(
    page: fitz.Page,
    api_key: str,
    model: str,
    session: requests.Session | None = None,
    timeout: float = 120.0,
) -> str:
    """Extract text from PDF page using SiliconFlow Multimodal OCR."""
    if not api_key:
        return ""

    pix = page.get_pixmap(dpi=200, alpha=False)
    img_data = pix.tobytes("png")
    base64_img = base64.b64encode(img_data).decode("utf-8")

    payload = {
        "model": model,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "请提取图中所有文字内容，按阅读顺序排列，保持原有段落结构，只需返回文字内容。"},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{base64_img}"},
                    },
                ],
            }
        ],
        "temperature": 0.1,
        "max_tokens": 2048,
    }
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    sess = session or requests.Session()
    try:
        # Use common SiliconFlow V1 endpoint
        response = sess.post(DEFAULT_SILICONFLOW_API_BASE, headers=headers, json=payload, timeout=timeout)
        if response.status_code == 200:
            data = response.json()
            return data["choices"][0]["message"]["content"]
        else:
            print(f"  [WARN] SiliconFlow OCR failed: {response.status_code} {response.text[:200]}", file=sys.stderr)
    except Exception as exc:
        print(f"  [WARN] SiliconFlow OCR exception: {exc}", file=sys.stderr)
    return ""


def ocr_pdf_page_llm(
    page: fitz.Page,
    api_key: str,
    model: str,
    api_base: str,
    session: requests.Session | None = None,
    timeout: float = 120.0,
) -> str:
    """Extract text from PDF page using the main LLM as fallback when SiliconFlow OCR fails."""
    if not api_key:
        return ""

    pix = page.get_pixmap(dpi=200, alpha=False)
    img_data = pix.tobytes("png")
    base64_img = base64.b64encode(img_data).decode("utf-8")

    payload = {
        "model": model,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "请提取图中所有文字内容，按阅读顺序排列，保持原有段落结构，只需返回文字内容。"},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{base64_img}"},
                    },
                ],
            }
        ],
        "temperature": 0.1,
        "max_tokens": 2048,
    }
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    sess = session or requests.Session()
    for attempt in range(3):
        try:
            response = sess.post(api_base, headers=headers, json=payload, timeout=timeout)
            if response.status_code == 200:
                data = response.json()
                return data["choices"][0]["message"]["content"]
            elif response.status_code in {429, 500, 502, 503, 504}:
                print(f"  [WARN] LLM OCR attempt {attempt+1} failed: {response.status_code} {response.text[:200]}", file=sys.stderr)
                time.sleep(min(2 ** attempt, 10))
                continue
            else:
                print(f"  [WARN] LLM OCR failed: {response.status_code} {response.text[:200]}", file=sys.stderr)
                return ""
        except Exception as exc:
            print(f"  [WARN] LLM OCR exception: {exc}", file=sys.stderr)
            time.sleep(min(2 ** attempt, 10))
    return ""


def is_low_quality_text(text: str, threshold: float = 0.2) -> bool:
    """Determine if extracted text is 'low quality' (e.g. mostly numbers/garbage)."""
    text = text.strip()
    if not text:
        return True
    
    # Heuristic: if digits far outnumber Chinese characters
    chinese_chars = len(re.findall(r"[\u4e00-\u9fff]", text))
    digits = len(re.findall(r"\d", text))
    
    if chinese_chars == 0 and digits > 5:
        return True
    
    if digits / (len(text) + 1) > threshold and chinese_chars < 5:
        return True
    
    return False


def extract_text_from_pdf(
    path: Path,
    ocr_strategy: str = "auto",
    ocr_api_key: str = "",
    ocr_model: str = DEFAULT_OCR_MODEL,
    ocr_threshold: float = DEFAULT_OCR_THRESHOLD,
    ocr_lang: str = "chi_sim+eng",
    llm_api_key: str = "",
    llm_model: str = DEFAULT_MODEL,
    llm_api_base: str = DEFAULT_API_BASE,
) -> list[tuple[str, str]]:
    """Extract layout-aware blocks from PDF with optional OCR fallback."""
    rows: list[tuple[str, str]] = []
    doc = fitz.open(path)
    session = requests.Session()
    try:
        for index, page in enumerate(doc, start=1):
            text_content = ""

            # 1. Attempt Native Extraction (unless forced OCR)
            if ocr_strategy != "siliconflow":
                blocks = page.get_text("blocks", sort=True)
                page_parts = [b[4].strip() for b in blocks if b[4].strip()]
                text_content = "\n\n".join(page_parts)
                text_content = normalize_whitespace(text_content)

            # 2. Trigger OCR Fallback or Force
            need_ocr = False
            if ocr_strategy == "siliconflow":
                need_ocr = True
            elif ocr_strategy == "auto" and is_low_quality_text(text_content, ocr_threshold):
                need_ocr = True
            elif not text_content:
                need_ocr = True

            if need_ocr:
                # Try SiliconFlow OCR first
                ocr_text = ocr_pdf_page_siliconflow(
                    page=page, api_key=ocr_api_key, model=ocr_model, session=session
                )
                # If SiliconFlow OCR fails, try LLM as fallback
                if not ocr_text and llm_api_key and llm_model:
                    print(f"  [INFO] SiliconFlow OCR failed for page {index}, trying LLM fallback...", file=sys.stderr)
                    ocr_text = ocr_pdf_page_llm(
                        page=page,
                        api_key=llm_api_key,
                        model=llm_model,
                        api_base=llm_api_base,
                        session=session,
                        timeout=120.0,
                    )
                if ocr_text:
                    text_content = normalize_whitespace(ocr_text)

            rows.append((str(index), text_content))
    finally:
        doc.close()
    return rows


def shutil_which(command: str) -> str | None:
    for directory in os.environ.get("PATH", "").split(os.pathsep):
        candidate = Path(directory) / command
        if candidate.is_file() and os.access(candidate, os.X_OK):
            return str(candidate)
    return None


def normalize_page_no(page_no: str) -> str:
    text = str(page_no).strip()
    if text.isdigit():
        return text
    match = re.search(r"\d+", text)
    if match:
        return match.group(0)
    return "1"


def extract_rows_for_file(path: Path, project_root: Path, args: argparse.Namespace) -> list[TextRow]:
    suffix = path.suffix.lower()
    if suffix == ".pptx":
        page_rows = extract_text_from_pptx(path)
    elif suffix == ".ppt":
        page_rows = extract_text_from_ppt(path)
    elif suffix == ".pdf":
        page_rows = extract_text_from_pdf(
            path=path,
            ocr_strategy=args.ocr_strategy,
            ocr_api_key=args.api_key if args.ocr_strategy in ("auto", "siliconflow") else "",
            ocr_model=args.ocr_model,
            ocr_threshold=args.ocr_threshold,
            llm_api_key=args.claude_api_key,
            llm_model=args.model,
            llm_api_base=DEFAULT_API_BASE,
        )
    elif suffix == ".docx":
        page_rows = extract_text_from_docx(path)
    elif suffix == ".doc":
        page_rows = extract_text_from_doc(path)
    else:
        return []

    try:
        relative_dir = str(path.parent.relative_to(project_root))
    except ValueError:
        relative_dir = str(path.parent)
    file_name = path.stem
    try:
        source_rel_path = str(path.relative_to(project_root))
    except ValueError:
        source_rel_path = str(path)

    rows: list[TextRow] = []
    for page_no, text in page_rows:
        page_no = normalize_page_no(page_no)
        rows.append(
            TextRow(
                text_id=f"{source_rel_path}#{page_no}",
                keywords="",
                directory=relative_dir,
                file_type=path.suffix.lstrip(".").lower(),
                file_name=file_name,
                page_no=page_no,
                text_content=text,
            )
        )
    return rows


def build_keyword_prompt(text: str, file_type: str, page_no: str) -> list[dict[str, str]]:
    clipped_text = text[:6000]
    return [
        {
            "role": "system",
            "content": (
                "你是投研资料编目助手。"
                "请从输入文字中提取 4 到 8 个中文关键词或短语，用于后续RAG混合检索。\n"
                "关键词生成规则：\n"
                "1. 每个关键词必须不少于6个中文字符，要包含定语修饰成分，例如“商业航天运载火箭”而不是“运载火箭”，“低轨卫星互联网星座”而不是“卫星”。\n"
                "2. 关键词应包含具体的技术名称、公司名称+业务描述、产品名称+应用场景、指标名称+度量对象等限定信息。\n"
                "3. 避免过于宽泛的短词，如“市场”“趋势”“概述”“发展”，必须带上修饰定语使其具备检索区分度。\n"
                "输出格式必须是用中文顿号“、”分隔的一行文本，不要编号，不要解释。"
            ),
        },
        {
            "role": "user",
            "content": (
                f"文件类型：{file_type}\n页码：{page_no}\n"
                "请提取关键词，每个关键词要长且包含定语，只返回关键词字符串。\n"
                f"文字内容：\n{clipped_text}"
            ),
        },
    ]


def request_keywords(
    session: requests.Session,
    limiter: RateLimiter,
    api_key: str,
    model: str,
    text: str,
    file_type: str,
    page_no: str,
    timeout: float,
) -> str:
    if not text.strip():
        return ""
    payload = {
        "model": model,
        "messages": build_keyword_prompt(text=text, file_type=file_type, page_no=page_no),
        "temperature": 0.2,
        "max_tokens": 256,
        "stream": False,
    }
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    for attempt in range(5):
        limiter.wait()
        response = session.post(DEFAULT_API_BASE, headers=headers, json=payload, timeout=timeout)
        if response.status_code == 200:
            data = response.json()
            content = data["choices"][0]["message"]["content"]
            return normalize_whitespace(str(content)).strip("，,;；")
        if response.status_code in {429, 500, 502, 503, 504}:
            time.sleep(min(2**attempt, 10))
            continue
        raise RuntimeError(f"SiliconFlow request failed with {response.status_code}: {response.text[:500]}")
    raise RuntimeError("SiliconFlow request failed after retries")


def enrich_keywords_concurrently(
    rows: list[TextRow],
    api_key: str,
    model: str,
    rpm: int,
    max_workers: int,
    timeout: float,
    force: bool = False,
) -> list[str]:
    limiter = RateLimiter(rpm=rpm)
    session_local = threading.local()
    failures: list[str] = []

    def task(row: TextRow) -> str:
        if not hasattr(session_local, "session"):
            session_local.session = requests.Session()
        return request_keywords(
            session=session_local.session,
            limiter=limiter,
            api_key=api_key,
            model=model,
            text=row.text_content,
            file_type=row.file_type,
            page_no=row.page_no,
            timeout=timeout,
        )

    targets = [row for row in rows if row.text_content.strip() and (force or not row.keywords.strip())]
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {executor.submit(task, row): row for row in targets}
        for future in concurrent.futures.as_completed(future_map):
            row = future_map[future]
            try:
                row.keywords = future.result()
            except Exception as exc:  # pragma: no cover
                failures.append(f"{row.text_id}: {exc}")
    return failures


def read_existing_text_rows(workbook_path: Path, sheet_name: str) -> dict[str, list[TextRow]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    rows_by_doc: dict[str, list[TextRow]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        text_id = str(row[0]).strip()
        text_row = TextRow(
            text_id=text_id,
            keywords=str(row[1] or "").strip(),
            directory=str(row[2] or "").strip(),
            file_type=str(row[3] or "").strip(),
            file_name=str(row[4] or "").strip(),
            page_no=str(row[5] or "").strip(),
            text_content=str(row[6] or ""),
        )
        rows_by_doc.setdefault(doc_key_from_text_id(text_id), []).append(text_row)
    return rows_by_doc


def write_rows_to_workbook(workbook_path: Path, sheet_name: str, rows: Sequence[TextRow]) -> None:
    wb = load_workbook(workbook_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    insert_index = 2 if len(wb.sheetnames) >= 2 else len(wb.sheetnames)
    ws = wb.create_sheet(title=sheet_name, index=insert_index)
    ws.append(("文本ID", "关键词", "目录", "文件类型", "文件名", "页码", "文本内容"))
    for row in rows:
        ws.append(
            (
                sanitize_for_excel(row.text_id),
                sanitize_for_excel(row.keywords),
                sanitize_for_excel(row.directory),
                sanitize_for_excel(row.file_type),
                sanitize_for_excel(row.file_name),
                sanitize_for_excel(row.page_no),
                sanitize_for_excel(row.text_content),
            )
        )
    wb.save(workbook_path)


def bootstrap_documents_from_sheet(
    workbook_state: dict,
    rows_by_doc: dict[str, list[TextRow]],
    project_root: Path,
) -> None:
    documents = workbook_state.setdefault("documents", {})
    for doc_key, rows in rows_by_doc.items():
        if doc_key in documents:
            continue
        candidate = Path(doc_key)
        if not candidate.is_absolute():
            candidate = (project_root / candidate).resolve()
        if candidate.exists() and candidate.suffix.lower() in SUPPORTED_SOURCE_SUFFIXES:
            documents[doc_key] = {
                "status": "done",
                "signature": file_signature(candidate),
                "row_count": len(rows),
                "updated_at": now_iso(),
                "bootstrap": True,
            }
        else:
            documents[doc_key] = {
                "status": "orphan",
                "row_count": len(rows),
                "updated_at": now_iso(),
                "bootstrap": True,
            }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Incremental text extraction into Excel worksheet #3 (`文本`).")
    parser.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT)
    parser.add_argument("--workbook", type=Path, default=None)
    parser.add_argument("--source-dir", type=Path, default=None)
    parser.add_argument("--project-root", type=Path, default=Path.cwd())
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--api-key", default="")
    parser.add_argument("--claude-api-key", default="")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--rpm", type=int, default=500)
    parser.add_argument("--max-workers", type=int, default=8)
    parser.add_argument("--timeout", type=float, default=120.0)
    parser.add_argument("--skip-keywords", action="store_true")
    parser.add_argument(
        "--bootstrap-workbooks",
        action="store_true",
        help="Auto-create missing annotation workbooks (标注.xlsx) before writing the third worksheet.",
    )
    parser.add_argument("--force-reextract", action="store_true", help="Re-extract all documents regardless of checkpoint.")
    parser.add_argument("--progress-file", type=Path, default=None)
    parser.add_argument("--ocr-strategy", default=DEFAULT_OCR_STRATEGY, choices=["auto", "fitz", "tesseract", "siliconflow"])
    parser.add_argument("--ocr-model", default=DEFAULT_OCR_MODEL)
    parser.add_argument("--ocr-threshold", type=float, default=DEFAULT_OCR_THRESHOLD)
    parser.add_argument("--ocr-lang", default="chi_sim+eng")
    args = parser.parse_args()
    args.api_key = resolve_env_value("SILICONFLOW_API_KEY", args.api_key)
    args.claude_api_key = resolve_claude_key()
    return args


def process_single_workbook(
    workbook_path: Path,
    source_dir: Path,
    args: argparse.Namespace,
    keyword_enabled: bool,
    progress: dict,
    progress_file: Path,
) -> tuple[int, list[str]]:
    workbook_key = str(workbook_path.resolve())
    workbook_state = progress["workbooks"].setdefault(
        workbook_key,
        {
            "workbook_path": workbook_key,
            "source_dir": str(source_dir.resolve()),
            "sheet_name": args.sheet_name,
            "status": "pending",
            "started_at": now_iso(),
            "updated_at": now_iso(),
            "summary": {"source_files": 0, "docs_skipped": 0, "docs_failed": 0, "rows_written": 0},
            "documents": {},
        },
    )
    workbook_state["status"] = "running"
    workbook_state["updated_at"] = now_iso()

    existing_rows_by_doc = read_existing_text_rows(workbook_path=workbook_path, sheet_name=args.sheet_name)
    bootstrap_documents_from_sheet(workbook_state=workbook_state, rows_by_doc=existing_rows_by_doc, project_root=args.project_root)
    save_progress(progress_file=progress_file, progress=progress)

    selected_rows_by_doc: dict[str, list[TextRow]] = {}
    failures: list[str] = []
    docs_skipped = 0

    source_files = list(iter_source_files(source_dir))
    workbook_state["summary"]["source_files"] = len(source_files)
    print(f"[Workbook] {workbook_path}")
    print(f"  Source dir: {source_dir}")
    print(f"  Found {len(source_files)} source files")

    documents = workbook_state.setdefault("documents", {})
    for path in source_files:
        try:
            try:
                doc_key = str(path.resolve().relative_to(args.project_root.resolve()))
            except ValueError:
                doc_key = str(path.resolve())
            signature = file_signature(path)
            doc_state = documents.get(doc_key, {})
            can_skip = (
                (not args.force_reextract)
                and doc_state.get("status") == "done"
                and doc_state.get("signature") == signature
                and doc_key in existing_rows_by_doc
            )
            if can_skip:
                selected_rows_by_doc[doc_key] = existing_rows_by_doc[doc_key]
                docs_skipped += 1
                doc_state["updated_at"] = now_iso()
                documents[doc_key] = doc_state
                continue

            rows = [row for row in extract_rows_for_file(path=path, project_root=args.project_root.resolve(), args=args) if row.text_content]
            selected_rows_by_doc[doc_key] = rows
            documents[doc_key] = {
                "status": "done",
                "signature": signature,
                "row_count": len(rows),
                "updated_at": now_iso(),
            }
            print(f"  [OK] {path.name}: {len(rows)} rows")
        except Exception as exc:  # pragma: no cover
            message = f"{path}: {exc}"
            failures.append(message)
            print(f"  [WARN] {message}", file=sys.stderr)
            try:
                doc_key = str(path.resolve().relative_to(args.project_root.resolve()))
            except ValueError:
                doc_key = str(path.resolve())
            documents[doc_key] = {
                "status": "failed",
                "last_error": str(exc),
                "updated_at": now_iso(),
            }
            if doc_key in existing_rows_by_doc:
                selected_rows_by_doc[doc_key] = existing_rows_by_doc[doc_key]
        workbook_state["updated_at"] = now_iso()
        workbook_state["summary"]["docs_skipped"] = docs_skipped
        workbook_state["summary"]["docs_failed"] = len(failures)
        save_progress(progress_file=progress_file, progress=progress)

    if not args.force_reextract:
        for doc_key, rows in existing_rows_by_doc.items():
            if doc_key not in selected_rows_by_doc:
                selected_rows_by_doc[doc_key] = rows
                documents.setdefault(
                    doc_key,
                    {"status": "orphan", "row_count": len(rows), "updated_at": now_iso()},
                )

    all_rows = sorted([row for rows in selected_rows_by_doc.values() for row in rows], key=row_sort_key)

    if keyword_enabled and all_rows:
        print(
            f"  Generating missing keywords for {len(all_rows)} rows with "
            f"{args.max_workers} workers / {args.rpm} RPM..."
        )
        kw_failures = enrich_keywords_concurrently(
            rows=all_rows,
            api_key=args.claude_api_key,
            model=args.model,
            rpm=args.rpm,
            max_workers=args.max_workers,
            timeout=args.timeout,
            force=False,
        )
        failures.extend(kw_failures)

    write_rows_to_workbook(workbook_path=workbook_path, sheet_name=args.sheet_name, rows=all_rows)
    print(f"  Wrote {len(all_rows)} rows to worksheet '{args.sheet_name}'")

    workbook_state["status"] = "completed" if not failures else "completed_with_failures"
    workbook_state["updated_at"] = now_iso()
    workbook_state["summary"]["docs_skipped"] = docs_skipped
    workbook_state["summary"]["docs_failed"] = len(failures)
    workbook_state["summary"]["rows_written"] = len(all_rows)
    save_progress(progress_file=progress_file, progress=progress)
    return len(all_rows), failures


def main() -> int:
    args = parse_args()
    args.project_root = args.project_root.resolve()
    data_root = args.data_root.resolve()
    progress_file = resolve_progress_file(data_root=data_root, progress_file=args.progress_file)
    progress = load_progress(progress_file=progress_file)

    jobs: list[WorkbookJob]
    if args.workbook:
        workbook_path = args.workbook.resolve()
        source_dir = args.source_dir.resolve() if args.source_dir else workbook_path.parent.resolve()
        if not source_dir.exists():
            raise FileNotFoundError(f"Source directory not found: {source_dir}")
        if not workbook_path.exists():
            if not args.bootstrap_workbooks:
                raise FileNotFoundError(f"Workbook not found: {workbook_path}")
            create_annotation_workbook(workbook_path)
        jobs = [WorkbookJob(workbook_path=workbook_path, source_dir=source_dir)]
    else:
        if not data_root.exists():
            raise FileNotFoundError(f"Data root not found: {data_root}")
        jobs = discover_or_bootstrap_workbook_jobs(data_root, bootstrap_workbooks=args.bootstrap_workbooks)
        if not jobs:
            raise FileNotFoundError(f"No annotation targets found under {data_root}")

    keyword_enabled = (not args.skip_keywords) and bool(args.api_key)
    if (not args.skip_keywords) and (not args.api_key):
        print("[WARN] Missing API key, keyword generation is skipped.", file=sys.stderr)

    print(f"Discovered {len(jobs)} workbook(s) to process.")
    total_rows = 0
    total_failures: list[str] = []
    for job in jobs:
        workbook_rows, failures = process_single_workbook(
            workbook_path=job.workbook_path.resolve(),
            source_dir=job.source_dir.resolve(),
            args=args,
            keyword_enabled=keyword_enabled,
            progress=progress,
            progress_file=progress_file,
        )
        total_rows += workbook_rows
        total_failures.extend(failures)

    print(f"\nAll done. Wrote {total_rows} rows across {len(jobs)} workbook(s).")
    if total_failures:
        print("\nWarnings / failures:", file=sys.stderr)
        for item in total_failures:
            print(f"- {item}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
