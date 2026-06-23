#!/usr/bin/env python3
"""Incrementally fill `关键词` column in worksheet `文本` with checkpoint resume."""

from __future__ import annotations

import argparse
import concurrent.futures
import hashlib
import json
import os
import threading
import time
from collections import deque
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

import requests
from openpyxl import load_workbook

from env_utils import resolve_env_value, resolve_dashscope_key


DEFAULT_DATA_ROOT = Path("source_data")
DEFAULT_PROGRESS_FILE_NAME = "extract_document_keywords_progress.json"
DEFAULT_SHEET = "文本"
DEFAULT_MODEL = "qwen3.6-plus"
DEFAULT_API_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
SUPPORTED_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
TEMP_PREFIXES = ("~$",)


@dataclass(frozen=True)
class WorkbookJob:
    workbook_path: Path


@dataclass(frozen=True)
class KeywordJob:
    row_idx: int
    row_key: str
    doc_key: str
    file_type: str
    page_no: str
    text: str
    text_hash: str


class RateLimiter:
    def __init__(self, rpm: int) -> None:
        if rpm <= 0:
            raise ValueError("rpm must be > 0")
        self.rpm = rpm
        self.window = deque()
        self.lock = threading.Lock()

    def wait(self) -> None:
        while True:
            with self.lock:
                now = time.monotonic()
                cutoff = now - 60.0
                while self.window and self.window[0] < cutoff:
                    self.window.popleft()
                if len(self.window) < self.rpm:
                    self.window.append(now)
                    return
                sleep_for = max(0.05, 60.0 - (now - self.window[0]))
            time.sleep(min(sleep_for, 1.0))


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize(text: str) -> str:
    return " ".join((text or "").replace("\r", "\n").split()).strip()


def load_progress(progress_file: Path) -> dict:
    if progress_file.exists():
        try:
            return json.loads(progress_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            pass
    return {"version": 1, "updated_at": now_iso(), "workbooks": {}}


def save_progress(progress_file: Path, progress: dict) -> None:
    progress["updated_at"] = now_iso()
    progress_file.parent.mkdir(parents=True, exist_ok=True)
    temp_path = progress_file.with_suffix(progress_file.suffix + ".tmp")
    temp_path.write_text(json.dumps(progress, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(progress_file)


def resolve_progress_file(data_root: Path, progress_file: Path | None) -> Path:
    if progress_file is not None:
        return progress_file.resolve()
    return (data_root / DEFAULT_PROGRESS_FILE_NAME).resolve()


def iter_workbooks(data_root: Path) -> Iterator[Path]:
    for path in sorted(data_root.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith(TEMP_PREFIXES):
            continue
        if path.suffix.lower() in SUPPORTED_WORKBOOK_SUFFIXES:
            yield path


def discover_workbooks(data_root: Path) -> list[WorkbookJob]:
    return [WorkbookJob(workbook_path=p) for p in iter_workbooks(data_root)]


def build_context_for_asset(
    ws: Any,
    row_idx: int,
    header: dict[str, int],
    data_root: Path,
    text_sheet_rows: dict[tuple[str, int], str],
) -> str:
    """Build text content for 图片/表格 rows from matched text or markdown content."""
    sheet_name = ws.title.strip()

    if sheet_name == "图片":
        # 图片: column A = image path, B = keyword (to fill)
        asset_path = str(ws.cell(row=row_idx, column=header.get("图片", 1)).value or "").strip()
        file_name = str(ws.cell(row=row_idx, column=header.get("文件名", 5)).value or "").strip()
        page_no = ws.cell(row=row_idx, column=header.get("页码", 6)).value
        page_no_val = int(page_no) if page_no and str(page_no).isdigit() else 0

        # Try to match to text sheet by (file_name, page_no)
        key = (file_name, page_no_val)
        matched_text = text_sheet_rows.get(key, "")
        if matched_text:
            return matched_text

        # Fallback: construct context from metadata
        return f"图片文件：{asset_path}\n来源文档：{file_name}\n页码：{page_no_val}"

    elif sheet_name == "表格":
        # 表格: column A = markdown file path
        md_path = str(ws.cell(row=row_idx, column=header.get("Markdown文件", 1)).value or "").strip()
        file_name = str(ws.cell(row=row_idx, column=header.get("文件名", 5)).value or "").strip()
        page_no = ws.cell(row=row_idx, column=header.get("页码", 6)).value
        page_no_val = int(page_no) if page_no and str(page_no).isdigit() else 0

        # Try to read markdown file
        md_full = data_root / md_path
        if md_full.exists():
            try:
                content = md_full.read_text(encoding="utf-8", errors="ignore")
                return content[:6000]
            except Exception:
                pass

        # Try to match to text sheet
        key = (file_name, page_no_val)
        matched_text = text_sheet_rows.get(key, "")
        if matched_text:
            return matched_text

        return f"表格文件：{md_path}\n来源文档：{file_name}\n页码：{page_no_val}"

    # For 文本 sheet or unknown sheets, caller provides text directly
    return ""


def build_messages(file_type: str, page_no: str, text: str) -> list[dict[str, str]]:
    return [
        {
            "role": "system",
            "content": (
                "你是投研资料编目助手。"
                "请从输入内容中提取 4 到 8 个中文关键词或短语，用于后续RAG混合检索。\n"
                "关键词生成规则：\n"
                "1. 每个关键词必须不少于6个中文字符，要包含定语修饰成分，例如“商业航天运载火箭”而不是“运载火箭”，“低轨卫星互联网星座”而不是“卫星”。\n"
                "2. 关键词应包含具体的技术名称、公司名称+业务描述、产品名称+应用场景、指标名称+度量对象等限定信息。\n"
                "3. 避免过于宽泛的短词，如“市场”“趋势”“概述”“发展”，必须带上修饰定语使其具备检索区分度。\n"
                "4. 输出必须是一行，使用中文顿号“、”分隔，不要解释，不要编号。"
            ),
        },
        {
            "role": "user",
            "content": (
                f"文件类型：{file_type}\n页码：{page_no}\n"
                "请提取关键词，每个关键词要长且包含定语，只返回关键词字符串。\n"
                f"文字内容：\n{text[:6000]}"
            ),
        },
    ]


def request_keywords(
    session: requests.Session,
    limiter: RateLimiter,
    api_key: str,
    model: str,
    file_type: str,
    page_no: str,
    text: str,
    timeout: float,
) -> str:
    payload = {
        "model": model,
        "messages": build_messages(file_type=file_type, page_no=page_no, text=text),
        "temperature": 0.2,
        "max_tokens": 256,
        "stream": False,
    }
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    for attempt in range(5):
        limiter.wait()
        response = session.post(DEFAULT_API_URL, headers=headers, json=payload, timeout=timeout)
        if response.status_code == 200:
            data = response.json()
            return normalize(data["choices"][0]["message"]["content"]).strip("，,;；")
        if response.status_code in {429, 500, 502, 503, 504}:
            time.sleep(min(2**attempt, 10))
            continue
        raise RuntimeError(f"SiliconFlow error {response.status_code}: {response.text[:500]}")
    raise RuntimeError("SiliconFlow request failed after retries")


def text_hash(text: str) -> str:
    return hashlib.sha1(text.encode("utf-8")).hexdigest()


def parse_doc_key(
    text_id: str,
    directory: str,
    file_type: str,
    file_name: str,
    row_idx: int,
) -> str:
    value = str(text_id or "").strip()
    if "#" in value:
        return value.rsplit("#", 1)[0]
    if directory or file_name or file_type:
        return f"{directory}/{file_name}.{file_type}".strip("/")
    return f"row:{row_idx}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Incremental keyword extraction for worksheet `文本`.")
    parser.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT)
    parser.add_argument("--workbook", type=Path, default=None, help="Optional single-workbook mode.")
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET)
    parser.add_argument("--api-key", default="")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--rpm", type=int, default=500)
    parser.add_argument("--max-workers", type=int, default=8)
    parser.add_argument("--timeout", type=float, default=120.0)
    parser.add_argument("--force", action="store_true", help="Regenerate keywords even when already filled.")
    parser.add_argument("--save-every", type=int, default=10, help="Save workbook and progress every N completions.")
    parser.add_argument("--progress-file", type=Path, default=None)
    args = parser.parse_args()
    args.api_key = resolve_dashscope_key(args.api_key)
    return args


def process_workbook(job: WorkbookJob, args: argparse.Namespace, progress: dict, progress_file: Path) -> tuple[int, int, int]:
    workbook_path = job.workbook_path.resolve()
    wb = load_workbook(workbook_path)
    if args.sheet_name not in wb.sheetnames:
        raise ValueError(f"Worksheet not found in {workbook_path}: {args.sheet_name}")
    ws = wb[args.sheet_name]

    header = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}
    required = ["关键词", "文件类型", "页码"]
    missing = [name for name in required if name not in header]
    if missing:
        raise ValueError(f"{workbook_path} missing columns: {missing}")

    keyword_col = header["关键词"]
    file_type_col = header["文件类型"]
    page_col = header["页码"]
    text_col = header.get("文本内容")
    text_id_col = header.get("文本ID")
    directory_col = header.get("目录")
    file_name_col = header.get("文件名")
    asset_col = header.get("图片") or header.get("Markdown文件") or header.get("文本ID")

    # Load text rows from 文本 sheet for context matching
    text_sheet_rows: dict[tuple[str, int], str] = {}
    if "文本" in wb.sheetnames:
        ws_text = wb["文本"]
        text_header = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws_text[1]) if cell.value}
        text_col_idx = text_header.get("文本内容")
        file_name_idx = text_header.get("文件名")
        page_idx = text_header.get("页码")
        if text_col_idx and file_name_idx and page_idx:
            for r in range(2, ws_text.max_row + 1):
                t = str(ws_text.cell(row=r, column=text_col_idx).value or "").strip()
                if t:
                    fn = str(ws_text.cell(row=r, column=file_name_idx).value or "").strip()
                    pn = ws_text.cell(row=r, column=page_idx).value
                    pn_val = int(pn) if pn and str(pn).isdigit() else 0
                    text_sheet_rows[(fn, pn_val)] = t


    workbook_key = str(workbook_path)
    workbook_state = progress["workbooks"].setdefault(
        workbook_key,
        {
            "workbook_path": workbook_key,
            "sheet_name": args.sheet_name,
            "status": "pending",
            "started_at": now_iso(),
            "updated_at": now_iso(),
            "summary": {"rows_total": 0, "rows_need": 0, "rows_completed": 0, "rows_failed": 0},
            "rows": {},
            "documents": {},
        },
    )
    workbook_state["status"] = "running"
    workbook_state["updated_at"] = now_iso()
    row_states: dict = workbook_state.setdefault("rows", {})

    # Bootstrap existing already-filled keywords into progress.
    for row_idx in range(2, ws.max_row + 1):
        # Build text content: for 文本 sheet use column directly, for 图片/表格 use context builder
        if text_col is not None:
            text = str(ws.cell(row=row_idx, column=text_col).value or "")
        else:
            text = build_context_for_asset(
                ws=ws,
                row_idx=row_idx,
                header=header,
                data_root=Path(args.data_root),
                text_sheet_rows=text_sheet_rows,
            )
        text_norm = normalize(text)
        if not text_norm:
            continue
        keyword = str(ws.cell(row=row_idx, column=keyword_col).value or "").strip()
        text_id = str(ws.cell(row=row_idx, column=text_id_col).value or "") if text_id_col else ""
        row_key = text_id if text_id else f"row:{row_idx}"
        doc_key = parse_doc_key(
            text_id=text_id,
            directory=str(ws.cell(row=row_idx, column=directory_col).value or "") if directory_col else "",
            file_type=str(ws.cell(row=row_idx, column=file_type_col).value or ""),
            file_name=str(ws.cell(row=row_idx, column=file_name_col).value or "") if file_name_col else "",
            row_idx=row_idx,
        )
        if keyword and (row_key not in row_states or row_states[row_key].get("status") != "done"):
            row_states[row_key] = {
                "status": "done",
                "row_index": row_idx,
                "doc_key": doc_key,
                "text_hash": text_hash(text_norm),
                "keyword": keyword,
                "updated_at": now_iso(),
                "bootstrap": True,
            }
    save_progress(progress_file=progress_file, progress=progress)

    jobs: list[KeywordJob] = []
    recovered = 0
    row_doc_map: dict[str, str] = {}
    rows_with_text: list[str] = []

    for row_idx in range(2, ws.max_row + 1):
        # Build text content: for 文本 sheet use column directly, for 图片/表格 use context builder
        if text_col is not None:
            text = str(ws.cell(row=row_idx, column=text_col).value or "")
        else:
            text = build_context_for_asset(
                ws=ws,
                row_idx=row_idx,
                header=header,
                data_root=Path(args.data_root),
                text_sheet_rows=text_sheet_rows,
            )
        text_norm = normalize(text)
        if not text_norm:
            continue
        text_id = str(ws.cell(row=row_idx, column=text_id_col).value or "") if text_id_col else ""
        row_key = text_id if text_id else f"row:{row_idx}"
        file_type = str(ws.cell(row=row_idx, column=file_type_col).value or "")
        page_no = str(ws.cell(row=row_idx, column=page_col).value or "")
        doc_key = parse_doc_key(
            text_id=text_id,
            directory=str(ws.cell(row=row_idx, column=directory_col).value or "") if directory_col else "",
            file_type=file_type,
            file_name=str(ws.cell(row=row_idx, column=file_name_col).value or "") if file_name_col else "",
            row_idx=row_idx,
        )
        row_doc_map[row_key] = doc_key
        rows_with_text.append(row_key)
        hash_value = text_hash(text_norm)
        keyword = str(ws.cell(row=row_idx, column=keyword_col).value or "").strip()
        state = row_states.get(row_key, {})

        if args.force:
            jobs.append(
                KeywordJob(
                    row_idx=row_idx,
                    row_key=row_key,
                    doc_key=doc_key,
                    file_type=file_type,
                    page_no=page_no,
                    text=text,
                    text_hash=hash_value,
                )
            )
            continue

        if keyword:
            continue

        if state.get("status") == "done" and state.get("text_hash") == hash_value and state.get("keyword"):
            ws.cell(row=row_idx, column=keyword_col, value=state["keyword"])
            recovered += 1
            continue

        jobs.append(
            KeywordJob(
                row_idx=row_idx,
                row_key=row_key,
                doc_key=doc_key,
                file_type=file_type,
                page_no=page_no,
                text=text,
                text_hash=hash_value,
            )
        )

    if recovered:
        wb.save(workbook_path)

    print(f"[Workbook] {workbook_path}")
    print(f"  Need keywords for {len(jobs)} rows")

    limiter = RateLimiter(args.rpm)
    local = threading.local()
    completed = 0
    failed = 0

    def task(job_item: KeywordJob) -> tuple[KeywordJob, str]:
        if not hasattr(local, "session"):
            local.session = requests.Session()
        keywords = request_keywords(
            session=local.session,
            limiter=limiter,
            api_key=args.api_key,
            model=args.model,
            file_type=job_item.file_type,
            page_no=job_item.page_no,
            text=job_item.text,
            timeout=args.timeout,
        )
        return job_item, keywords

    with concurrent.futures.ThreadPoolExecutor(max_workers=args.max_workers) as executor:
        future_map = {executor.submit(task, job_item): job_item for job_item in jobs}
        for future in concurrent.futures.as_completed(future_map):
            job_item = future_map[future]
            try:
                _, keywords = future.result()
                ws.cell(row=job_item.row_idx, column=keyword_col, value=keywords)
                row_states[job_item.row_key] = {
                    "status": "done",
                    "row_index": job_item.row_idx,
                    "doc_key": job_item.doc_key,
                    "text_hash": job_item.text_hash,
                    "keyword": keywords,
                    "updated_at": now_iso(),
                }
                completed += 1
                if completed % 5 == 0 or completed == len(jobs):
                    print(f"  Completed {completed}/{len(jobs)}")
            except Exception as exc:  # pragma: no cover
                failed += 1
                row_states[job_item.row_key] = {
                    "status": "failed",
                    "row_index": job_item.row_idx,
                    "doc_key": job_item.doc_key,
                    "text_hash": job_item.text_hash,
                    "error": str(exc),
                    "updated_at": now_iso(),
                }
                print(f"  [WARN] row {job_item.row_idx}: {exc}")
            if args.save_every > 0 and (completed + failed) % args.save_every == 0:
                wb.save(workbook_path)
                workbook_state["updated_at"] = now_iso()
                workbook_state["summary"]["rows_completed"] = completed
                workbook_state["summary"]["rows_failed"] = failed
                save_progress(progress_file=progress_file, progress=progress)

    wb.save(workbook_path)

    # Build document-level completion status.
    doc_totals: dict[str, int] = {}
    doc_done: dict[str, int] = {}
    for row_key in rows_with_text:
        doc_key = row_doc_map.get(row_key, "unknown")
        doc_totals[doc_key] = doc_totals.get(doc_key, 0) + 1
        state = row_states.get(row_key, {})
        if state.get("status") == "done" and str(state.get("keyword", "")).strip():
            doc_done[doc_key] = doc_done.get(doc_key, 0) + 1
    documents = workbook_state.setdefault("documents", {})
    for doc_key, total in doc_totals.items():
        done_count = doc_done.get(doc_key, 0)
        documents[doc_key] = {
            "status": "done" if done_count == total else "incomplete",
            "row_total": total,
            "row_done": done_count,
            "updated_at": now_iso(),
        }

    workbook_state["status"] = "completed" if failed == 0 else "completed_with_failures"
    workbook_state["updated_at"] = now_iso()
    workbook_state["summary"]["rows_total"] = len(rows_with_text)
    workbook_state["summary"]["rows_need"] = len(jobs)
    workbook_state["summary"]["rows_completed"] = completed
    workbook_state["summary"]["rows_failed"] = failed
    save_progress(progress_file=progress_file, progress=progress)
    return len(rows_with_text), completed, failed


def main() -> int:
    args = parse_args()
    if not args.api_key:
        raise ValueError("Missing SiliconFlow API key. Use --api-key or SILICONFLOW_API_KEY.")

    data_root = args.data_root.resolve()
    progress_file = resolve_progress_file(data_root=data_root, progress_file=args.progress_file)
    progress = load_progress(progress_file=progress_file)

    if args.workbook:
        jobs = [WorkbookJob(workbook_path=args.workbook.resolve())]
    else:
        if not data_root.exists():
            raise FileNotFoundError(f"Data root not found: {data_root}")
        jobs = discover_workbooks(data_root=data_root)
        if not jobs:
            raise FileNotFoundError(f"No Excel files found under {data_root}")

    total_rows = 0
    total_completed = 0
    total_failed = 0
    for job in jobs:
        if not job.workbook_path.exists():
            print(f"[WARN] Workbook not found, skip: {job.workbook_path}")
            continue
        rows_total, rows_completed, rows_failed = process_workbook(
            job=job, args=args, progress=progress, progress_file=progress_file
        )
        total_rows += rows_total
        total_completed += rows_completed
        total_failed += rows_failed

    print(
        f"\nAll done. Rows with text: {total_rows}, "
        f"newly completed keywords: {total_completed}, failed: {total_failed}."
    )
    return 0 if total_failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
